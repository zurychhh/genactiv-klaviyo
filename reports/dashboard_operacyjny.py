#!/usr/bin/env python3
"""
Dashboard Operacyjny — GenActiv.pl
Generuje Excel z danymi operacyjnymi z wszystkich zrodel.

Uzycie:
  source venv/bin/activate
  python3 reports/dashboard_operacyjny.py

Dane Shopify pobierane sa LIVE z API (GraphQL).
Dane GA4/Klaviyo/Meta/Google Ads — wklej z MCP lub uruchom skrypt uzupelniajacy.
Szczegoly w arkuszu "Instrukcja MCP".
"""

import json
import sys
import os
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime, timedelta
from collections import defaultdict, Counter
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

# ============================================================
# CONFIGURATION
# ============================================================
SHOPIFY_DOMAIN = os.environ.get("SHOPIFY_DOMAIN", "genactiv.myshopify.com")
SHOPIFY_TOKEN = os.environ.get("SHOPIFY_ACCESS_TOKEN", "")
SHOPIFY_API_VERSION = "2025-01"
SHOPIFY_ENDPOINT = f"https://{SHOPIFY_DOMAIN}/admin/api/{SHOPIFY_API_VERSION}/graphql.json"
SHOPIFY_HEADERS = {
    "Content-Type": "application/json",
    "X-Shopify-Access-Token": SHOPIFY_TOKEN,
}

# Periods
today = datetime.now()
period_end = today - timedelta(days=1)  # yesterday
period_start = period_end - timedelta(days=6)  # 7 days
period_prev_end = period_start - timedelta(days=1)
period_prev_start = period_prev_end - timedelta(days=6)

DATE_FMT = "%Y-%m-%d"
DATE_FMT_PL = "%d.%m.%Y"

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(
    OUTPUT_DIR, f"dashboard_operacyjny_{today.strftime(DATE_FMT)}.xlsx"
)

# ============================================================
# STYLES (matching existing GenActiv report style)
# ============================================================
NAVY = "1A3B5D"
BLUE = "4472C4"
RED = "EF3340"
GREEN = "27AE60"
SHOPIFY_GREEN = "95BF47"

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color=NAVY)
TITLE_FONT = Font(name="Calibri", bold=True, size=16, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=10, color="666666")
KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=14, color=NAVY)
KPI_LABEL_FONT = Font(name="Calibri", size=9, color="666666")
BOLD = Font(name="Calibri", bold=True, size=10)
NORMAL = Font(name="Calibri", size=10)
PLACEHOLDER_FONT = Font(name="Calibri", size=10, color="999999", italic=True)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

PLN_FORMAT = '#,##0" PLN"'
PCT_FORMAT = "0.0%"
NUM_FORMAT = "#,##0"


# ============================================================
# HELPER FUNCTIONS — STYLING
# ============================================================
def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_section_row(ws, row, max_col, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 28
    return row


def apply_borders(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def auto_width(ws, max_col, min_width=10, max_width=40):
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def write_placeholder(ws, row, col, comment_text=None):
    """Write a placeholder dash with optional comment."""
    cell = ws.cell(row=row, column=col, value="\u2014")
    cell.font = PLACEHOLDER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if comment_text:
        cell.comment = Comment(comment_text, "Dashboard Generator")
    return cell


def delta_pct(current, previous):
    """Calculate percentage change, return None if not possible."""
    if previous and previous != 0:
        return (current - previous) / previous
    return None


def fmt_delta(value):
    """Format delta as string with arrow."""
    if value is None:
        return "n/d"
    sign = "+" if value >= 0 else ""
    return f"{sign}{value:.1%}"


# ============================================================
# SHOPIFY GRAPHQL API
# ============================================================
def shopify_graphql(query, variables=None):
    """Execute a Shopify GraphQL query with error handling."""
    payload = {"query": query}
    if variables:
        payload["variables"] = variables
    try:
        resp = requests.post(SHOPIFY_ENDPOINT, headers=SHOPIFY_HEADERS, json=payload, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        if "errors" in data:
            print(f"  [WARN] GraphQL errors: {data['errors']}")
        return data
    except requests.RequestException as e:
        print(f"  [ERROR] Shopify API: {e}")
        return None


def get_shopify_orders(date_from, date_to):
    """Fetch ALL orders in date range with pagination."""
    all_orders = []
    cursor = None
    page = 0

    query_template = """
    query($cursor: String) {
      orders(first: 50, after: $cursor, sortKey: CREATED_AT, reverse: true,
             query: "created_at:>=%s created_at:<=%s") {
        edges {
          cursor
          node {
            name
            createdAt
            totalPriceSet { shopMoney { amount currencyCode } }
            subtotalPriceSet { shopMoney { amount } }
            totalDiscountsSet { shopMoney { amount } }
            displayFinancialStatus
            displayFulfillmentStatus
            customer {
              email
              firstName
              lastName
              numberOfOrders
            }
            customerJourneySummary {
              firstVisit {
                source
                sourceType
                utmParameters { source medium campaign }
              }
              lastVisit {
                source
                sourceType
                utmParameters { source medium campaign }
              }
            }
            lineItems(first: 20) {
              edges {
                node {
                  title
                  quantity
                  originalTotalSet { shopMoney { amount } }
                }
              }
            }
          }
        }
        pageInfo { hasNextPage endCursor }
      }
    }
    """
    query = query_template % (date_from, date_to)

    while True:
        page += 1
        print(f"    Strona {page}... ", end="", flush=True)
        variables = {"cursor": cursor} if cursor else {}
        result = shopify_graphql(query, variables)

        if not result or "data" not in result:
            print("BLAD API")
            break

        edges = result["data"]["orders"]["edges"]
        print(f"{len(edges)} zamowien")
        for edge in edges:
            all_orders.append(edge["node"])

        page_info = result["data"]["orders"]["pageInfo"]
        if page_info["hasNextPage"]:
            cursor = edges[-1]["cursor"]
        else:
            break

    return all_orders


def analyze_orders(orders):
    """Analyze list of orders and return summary dict."""
    summary = {
        "count": len(orders),
        "revenue": 0.0,
        "avg_order_value": 0.0,
        "new_customers": 0,
        "returning_customers": 0,
        "by_day": defaultdict(lambda: {"count": 0, "revenue": 0.0}),
        "products": Counter(),
        "sources": Counter(),
        "financial_status": Counter(),
        "fulfillment_status": Counter(),
    }

    for order in orders:
        total = float(order["totalPriceSet"]["shopMoney"]["amount"])
        summary["revenue"] += total

        # Date breakdown
        created = order["createdAt"][:10]
        summary["by_day"][created]["count"] += 1
        summary["by_day"][created]["revenue"] += total

        # Customer type
        customer = order.get("customer")
        if customer:
            try:
                orders_count = int(customer.get("numberOfOrders", 0) or 0)
            except (ValueError, TypeError):
                orders_count = 0
            if orders_count > 1:
                summary["returning_customers"] += 1
            else:
                summary["new_customers"] += 1
        else:
            summary["new_customers"] += 1

        # Products
        for li_edge in order.get("lineItems", {}).get("edges", []):
            li = li_edge["node"]
            summary["products"][li["title"]] += li["quantity"]

        # Traffic sources (last visit)
        journey = order.get("customerJourneySummary")
        if journey and journey.get("lastVisit"):
            visit = journey["lastVisit"]
            source = visit.get("source") or "(unknown)"
            utm = visit.get("utmParameters")
            if utm and utm.get("source"):
                source_label = f"{utm['source']} / {utm.get('medium', '?')}"
            else:
                source_label = source
            summary["sources"][source_label] += 1
        else:
            summary["sources"]["(brak danych journey)"] += 1

        # Status
        summary["financial_status"][order.get("displayFinancialStatus", "?")] += 1
        summary["fulfillment_status"][order.get("displayFulfillmentStatus", "?")] += 1

    if summary["count"] > 0:
        summary["avg_order_value"] = summary["revenue"] / summary["count"]

    return summary


# ============================================================
# MAIN — BUILD EXCEL
# ============================================================
def main():
    print("=" * 60)
    print("DASHBOARD OPERACYJNY — GenActiv.pl")
    print("=" * 60)
    print(f"Data generowania: {today.strftime('%Y-%m-%d %H:%M')}")
    print(f"Okres biezacy:    {period_start.strftime(DATE_FMT)} — {period_end.strftime(DATE_FMT)}")
    print(f"Okres poprzedni:  {period_prev_start.strftime(DATE_FMT)} — {period_prev_end.strftime(DATE_FMT)}")
    print()

    # ----------------------------------------------------------
    # FETCH SHOPIFY DATA
    # ----------------------------------------------------------
    print("[1/4] Pobieranie zamowien Shopify (biezacy okres)...")
    orders_current = get_shopify_orders(
        period_start.strftime(DATE_FMT), period_end.strftime(DATE_FMT)
    )
    print(f"  Pobrano {len(orders_current)} zamowien\n")

    print("[2/4] Pobieranie zamowien Shopify (poprzedni okres)...")
    orders_prev = get_shopify_orders(
        period_prev_start.strftime(DATE_FMT), period_prev_end.strftime(DATE_FMT)
    )
    print(f"  Pobrano {len(orders_prev)} zamowien\n")

    print("[3/4] Analiza danych...")
    curr = analyze_orders(orders_current)
    prev = analyze_orders(orders_prev)
    print(f"  Biezacy:    {curr['count']} zamowien, {curr['revenue']:.0f} PLN")
    print(f"  Poprzedni:  {prev['count']} zamowien, {prev['revenue']:.0f} PLN")
    print()

    print("[4/4] Generowanie Excel...")
    wb = openpyxl.Workbook()

    # ==========================================================
    # SHEET 1: DASHBOARD
    # ==========================================================
    ws1 = wb.active
    ws1.title = "Dashboard"
    ws1.sheet_properties.tabColor = NAVY
    MAX_COL_DASH = 8

    # Title
    ws1.merge_cells("A1:H1")
    ws1.cell(row=1, column=1, value="DASHBOARD OPERACYJNY \u2014 GenActiv.pl").font = TITLE_FONT
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:H2")
    ws1.cell(
        row=2, column=1,
        value=f"Wygenerowano: {today.strftime('%Y-%m-%d %H:%M')}  |  "
              f"Okres: {period_start.strftime(DATE_FMT_PL)} \u2013 {period_end.strftime(DATE_FMT_PL)}  |  "
              f"Poprzedni: {period_prev_start.strftime(DATE_FMT_PL)} \u2013 {period_prev_end.strftime(DATE_FMT_PL)}"
    ).font = SUBTITLE_FONT

    # ----------------------------------------------------------
    # Section: SPRZEDAZ (LIVE)
    # ----------------------------------------------------------
    row = 4
    style_section_row(ws1, row, MAX_COL_DASH, "SPRZEDAZ (Shopify \u2014 LIVE)")

    # KPI header row
    row += 1
    kpi_headers = ["KPI", "Biezacy 7d", "Poprzedni 7d", "\u0394 %"]
    for i, h in enumerate(kpi_headers):
        col = i + 2  # start from B
        cell = ws1.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # KPI data rows
    kpi_data = [
        (
            "Zamowienia",
            curr["count"],
            prev["count"],
            delta_pct(curr["count"], prev["count"]),
            NUM_FORMAT,
        ),
        (
            "Przychod PLN",
            curr["revenue"],
            prev["revenue"],
            delta_pct(curr["revenue"], prev["revenue"]),
            PLN_FORMAT,
        ),
        (
            "Srednia wartosc zamowienia (AOV)",
            curr["avg_order_value"],
            prev["avg_order_value"],
            delta_pct(curr["avg_order_value"], prev["avg_order_value"]),
            PLN_FORMAT,
        ),
        (
            "Nowi klienci",
            curr["new_customers"],
            prev["new_customers"],
            delta_pct(curr["new_customers"], prev["new_customers"]),
            NUM_FORMAT,
        ),
        (
            "Powracajacy klienci",
            curr["returning_customers"],
            prev["returning_customers"],
            delta_pct(curr["returning_customers"], prev["returning_customers"]),
            NUM_FORMAT,
        ),
    ]

    for label, val_curr, val_prev, delta, fmt in kpi_data:
        row += 1
        ws1.cell(row=row, column=2, value=label).font = BOLD
        cell_c = ws1.cell(row=row, column=3, value=val_curr)
        cell_c.font = KPI_VALUE_FONT
        cell_c.number_format = fmt
        cell_c.alignment = Alignment(horizontal="center")

        cell_p = ws1.cell(row=row, column=4, value=val_prev)
        cell_p.font = NORMAL
        cell_p.number_format = fmt
        cell_p.alignment = Alignment(horizontal="center")

        cell_d = ws1.cell(row=row, column=5, value=delta if delta is not None else "n/d")
        if isinstance(delta, (int, float)):
            cell_d.number_format = PCT_FORMAT
            if delta > 0:
                cell_d.fill = GREEN_FILL
                cell_d.font = Font(name="Calibri", bold=True, size=10, color="006100")
            elif delta < 0:
                cell_d.fill = RED_FILL
                cell_d.font = Font(name="Calibri", bold=True, size=10, color="9C0006")
            else:
                cell_d.font = NORMAL
        else:
            cell_d.font = PLACEHOLDER_FONT
        cell_d.alignment = Alignment(horizontal="center")

    apply_borders(ws1, 5, row, 5)

    # ----------------------------------------------------------
    # Section: RUCH (GA4 — placeholder)
    # ----------------------------------------------------------
    row += 2
    style_section_row(ws1, row, MAX_COL_DASH, "RUCH (GA4 \u2014 do uzupelnienia przez MCP)")

    row += 1
    ga4_headers = [
        "Kanal", "Sesje biezacy 7d", "Sesje poprzedni 7d",
        "\u0394 %", "Konwersje", "CR%", "Przychod PLN"
    ]
    for i, h in enumerate(ga4_headers):
        col = i + 1
        cell = ws1.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[row].height = 30
    ga4_header_row = row

    ga4_channels = [
        "Organic Search",
        "Paid Search (Google)",
        "Paid Search (Bing)",
        "Paid Social",
        "Organic Social",
        "Email",
        "Direct",
        "Referral",
        "Unassigned",
        "SUMA",
    ]
    mcp_comment = (
        "Uzupelnij: mcp__ga4__run_report\n"
        "property_id: 279858535\n"
        "dimensions: [\"sessionDefaultChannelGroup\"]\n"
        "metrics: [\"sessions\", \"conversions\", \"totalRevenue\"]"
    )

    for channel in ga4_channels:
        row += 1
        ws1.cell(row=row, column=1, value=channel).font = BOLD if channel == "SUMA" else NORMAL
        for c in range(2, len(ga4_headers) + 1):
            write_placeholder(ws1, row, c, mcp_comment)

    apply_borders(ws1, ga4_header_row, row, len(ga4_headers))

    # ----------------------------------------------------------
    # Section: EMAIL (Klaviyo — placeholder)
    # ----------------------------------------------------------
    row += 2
    style_section_row(ws1, row, MAX_COL_DASH, "EMAIL (Klaviyo \u2014 do uzupelnienia przez MCP)")

    row += 1
    email_headers = [
        "Kampania", "Data wysylki", "Wyslano", "Opened",
        "OR%", "Clicked", "CTR%", "Revenue PLN"
    ]
    for i, h in enumerate(email_headers):
        col = i + 1
        cell = ws1.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[row].height = 30
    email_header_row = row

    klaviyo_comment = (
        "Uzupelnij:\n"
        "1. mcp__klaviyo__klaviyo_get_campaigns (channel: email)\n"
        "2. mcp__klaviyo__klaviyo_get_campaign_report\n"
        "   (wymaga conversion_metric_id — uzyj mcp__klaviyo__klaviyo_get_metrics)"
    )

    for i in range(5):  # 5 placeholder rows
        row += 1
        for c in range(1, len(email_headers) + 1):
            write_placeholder(ws1, row, c, klaviyo_comment)

    apply_borders(ws1, email_header_row, row, len(email_headers))

    # ----------------------------------------------------------
    # Section: PAID ADS (Google + Meta — placeholder)
    # ----------------------------------------------------------
    row += 2
    style_section_row(ws1, row, MAX_COL_DASH, "PAID ADS (Google + Meta \u2014 do uzupelnienia przez MCP)")

    row += 1
    ads_headers = [
        "Platforma", "Kampania", "Wydatek PLN", "Klikniecia",
        "CPC PLN", "Konwersje", "CPA PLN", "ROAS"
    ]
    for i, h in enumerate(ads_headers):
        col = i + 1
        cell = ws1.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[row].height = 30
    ads_header_row = row

    ads_rows = [
        ("Google Ads", "Performance Max"),
        ("Google Ads", "Remarketing"),
        ("Google Ads", "Produktowa"),
        ("Bing Ads", "Sembot - wszystkie"),
        ("Meta", "Instagram - paid"),
        ("Meta", "Facebook - paid"),
        ("TikTok", "ZimazColostrum"),
        ("", "SUMA"),
    ]

    google_ads_comment = (
        "Uzupelnij: mcp__google-ads__run_gaql\n"
        "customer_id: 3393382047\n"
        "query: SELECT campaign.name, metrics.impressions, metrics.clicks,\n"
        "  metrics.cost_micros, metrics.conversions, metrics.conversions_value\n"
        "  FROM campaign WHERE segments.date DURING LAST_7_DAYS\n"
        "  ORDER BY metrics.cost_micros DESC"
    )
    meta_comment = (
        "Uzupelnij: mcp__meta-ads__get_insights\n"
        "lub mcp__meta-ads__get_campaigns + get_insights per campaign"
    )

    for platform, campaign in ads_rows:
        row += 1
        ws1.cell(row=row, column=1, value=platform).font = BOLD if campaign == "SUMA" else NORMAL
        ws1.cell(row=row, column=2, value=campaign).font = BOLD if campaign == "SUMA" else NORMAL
        comment = google_ads_comment if "Google" in platform or "Bing" in platform else meta_comment
        if campaign == "SUMA":
            comment = "Suma automatyczna po uzupelnieniu danych"
        for c in range(3, len(ads_headers) + 1):
            write_placeholder(ws1, row, c, comment)

    apply_borders(ws1, ads_header_row, row, len(ads_headers))

    # Set column widths for Dashboard
    col_widths_dash = {1: 30, 2: 22, 3: 22, 4: 22, 5: 14, 6: 14, 7: 14, 8: 14}
    for c, w in col_widths_dash.items():
        ws1.column_dimensions[get_column_letter(c)].width = w
    ws1.freeze_panes = "A4"

    # ==========================================================
    # SHEET 2: SHOPIFY SZCZEGOLY
    # ==========================================================
    print("  Arkusz: Shopify szczegoly...")
    ws2 = wb.create_sheet("Shopify szczegoly")
    ws2.sheet_properties.tabColor = SHOPIFY_GREEN
    MAX_COL_S2 = 7

    ws2.merge_cells("A1:G1")
    ws2.cell(
        row=1, column=1,
        value=f"SHOPIFY SZCZEGOLY \u2014 {period_start.strftime(DATE_FMT_PL)} \u2013 {period_end.strftime(DATE_FMT_PL)}"
    ).font = TITLE_FONT
    ws2.merge_cells("A2:G2")
    ws2.cell(
        row=2, column=1,
        value=f"Dane pobrane LIVE z Shopify GraphQL API  |  Zamowien: {curr['count']}"
    ).font = SUBTITLE_FONT

    # --- Orders by day ---
    row2 = 4
    style_section_row(ws2, row2, MAX_COL_S2, "ZAMOWIENIA WG DNIA")
    row2 += 1
    day_headers = ["Data", "Zamowienia", "Przychod PLN", "AOV PLN"]
    for i, h in enumerate(day_headers):
        ws2.cell(row=row2, column=i + 1, value=h)
    style_header_row(ws2, row2, len(day_headers))
    day_header_row = row2

    sorted_days = sorted(curr["by_day"].keys())
    total_orders_check = 0
    total_rev_check = 0.0
    for day in sorted_days:
        row2 += 1
        d = curr["by_day"][day]
        aov = d["revenue"] / d["count"] if d["count"] > 0 else 0
        ws2.cell(row=row2, column=1, value=day).font = NORMAL
        ws2.cell(row=row2, column=2, value=d["count"]).font = NORMAL
        ws2.cell(row=row2, column=2).alignment = Alignment(horizontal="center")
        c_rev = ws2.cell(row=row2, column=3, value=d["revenue"])
        c_rev.font = NORMAL
        c_rev.number_format = PLN_FORMAT
        c_aov = ws2.cell(row=row2, column=4, value=aov)
        c_aov.font = NORMAL
        c_aov.number_format = PLN_FORMAT
        total_orders_check += d["count"]
        total_rev_check += d["revenue"]

    # Totals
    row2 += 1
    ws2.cell(row=row2, column=1, value="SUMA").font = BOLD
    ws2.cell(row=row2, column=2, value=total_orders_check).font = BOLD
    ws2.cell(row=row2, column=2).alignment = Alignment(horizontal="center")
    c_tot = ws2.cell(row=row2, column=3, value=total_rev_check)
    c_tot.font = BOLD
    c_tot.number_format = PLN_FORMAT
    c_aov_tot = ws2.cell(
        row=row2, column=4,
        value=total_rev_check / total_orders_check if total_orders_check > 0 else 0
    )
    c_aov_tot.font = BOLD
    c_aov_tot.number_format = PLN_FORMAT

    apply_borders(ws2, day_header_row, row2, len(day_headers))

    # --- Top products ---
    row2 += 2
    style_section_row(ws2, row2, MAX_COL_S2, "TOP PRODUKTY (wg ilosci sprzedanej)")
    row2 += 1
    prod_headers = ["Produkt", "Ilosc"]
    for i, h in enumerate(prod_headers):
        ws2.cell(row=row2, column=i + 1, value=h)
    style_header_row(ws2, row2, len(prod_headers))
    prod_header_row = row2

    top_products = curr["products"].most_common(15)
    for title, qty in top_products:
        row2 += 1
        ws2.cell(row=row2, column=1, value=title).font = NORMAL
        ws2.cell(row=row2, column=2, value=qty).font = NORMAL
        ws2.cell(row=row2, column=2).alignment = Alignment(horizontal="center")

    apply_borders(ws2, prod_header_row, row2, len(prod_headers))

    # --- Traffic sources ---
    row2 += 2
    style_section_row(ws2, row2, MAX_COL_S2, "ZRODLA RUCHU (Shopify Customer Journey \u2014 last visit)")
    row2 += 1
    src_headers = ["Zrodlo / medium", "Zamowienia", "% udzialu"]
    for i, h in enumerate(src_headers):
        ws2.cell(row=row2, column=i + 1, value=h)
    style_header_row(ws2, row2, len(src_headers))
    src_header_row = row2

    sorted_sources = curr["sources"].most_common(20)
    total_src = sum(v for _, v in sorted_sources)
    for src, count in sorted_sources:
        row2 += 1
        ws2.cell(row=row2, column=1, value=src).font = NORMAL
        ws2.cell(row=row2, column=2, value=count).font = NORMAL
        ws2.cell(row=row2, column=2).alignment = Alignment(horizontal="center")
        share = count / total_src if total_src > 0 else 0
        cell_share = ws2.cell(row=row2, column=3, value=share)
        cell_share.font = NORMAL
        cell_share.number_format = PCT_FORMAT
        cell_share.alignment = Alignment(horizontal="center")

    apply_borders(ws2, src_header_row, row2, len(src_headers))

    # --- Customer type ---
    row2 += 2
    style_section_row(ws2, row2, MAX_COL_S2, "NOWI vs POWRACAJACY KLIENCI")
    row2 += 1
    cust_headers = ["Typ klienta", "Zamowienia", "% udzialu"]
    for i, h in enumerate(cust_headers):
        ws2.cell(row=row2, column=i + 1, value=h)
    style_header_row(ws2, row2, len(cust_headers))
    cust_header_row = row2

    for label, count in [
        ("Nowi (1. zamowienie)", curr["new_customers"]),
        ("Powracajacy (2+ zamowien)", curr["returning_customers"]),
    ]:
        row2 += 1
        ws2.cell(row=row2, column=1, value=label).font = NORMAL
        ws2.cell(row=row2, column=2, value=count).font = NORMAL
        ws2.cell(row=row2, column=2).alignment = Alignment(horizontal="center")
        share = count / curr["count"] if curr["count"] > 0 else 0
        cell_s = ws2.cell(row=row2, column=3, value=share)
        cell_s.font = NORMAL
        cell_s.number_format = PCT_FORMAT
        cell_s.alignment = Alignment(horizontal="center")

    apply_borders(ws2, cust_header_row, row2, len(cust_headers))

    # --- Payment status ---
    row2 += 2
    style_section_row(ws2, row2, MAX_COL_S2, "STATUS PLATNOSCI")
    row2 += 1
    pay_headers = ["Status", "Zamowienia"]
    for i, h in enumerate(pay_headers):
        ws2.cell(row=row2, column=i + 1, value=h)
    style_header_row(ws2, row2, len(pay_headers))
    pay_header_row = row2

    for status, count in curr["financial_status"].most_common():
        row2 += 1
        ws2.cell(row=row2, column=1, value=status).font = NORMAL
        ws2.cell(row=row2, column=2, value=count).font = NORMAL
        ws2.cell(row=row2, column=2).alignment = Alignment(horizontal="center")

    apply_borders(ws2, pay_header_row, row2, len(pay_headers))

    # --- Fulfillment status ---
    row2 += 2
    style_section_row(ws2, row2, MAX_COL_S2, "STATUS REALIZACJI")
    row2 += 1
    ful_headers = ["Status", "Zamowienia"]
    for i, h in enumerate(ful_headers):
        ws2.cell(row=row2, column=i + 1, value=h)
    style_header_row(ws2, row2, len(ful_headers))
    ful_header_row = row2

    for status, count in curr["fulfillment_status"].most_common():
        row2 += 1
        ws2.cell(row=row2, column=1, value=status).font = NORMAL
        ws2.cell(row=row2, column=2, value=count).font = NORMAL
        ws2.cell(row=row2, column=2).alignment = Alignment(horizontal="center")

    apply_borders(ws2, ful_header_row, row2, len(ful_headers))

    # Column widths
    col_widths_s2 = {1: 45, 2: 16, 3: 16, 4: 16, 5: 14, 6: 14, 7: 14}
    for c, w in col_widths_s2.items():
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.freeze_panes = "A4"

    # ==========================================================
    # SHEET 3: TRENDY TYGODNIOWE
    # ==========================================================
    print("  Arkusz: Trendy tygodniowe...")
    ws3 = wb.create_sheet("Trendy tygodniowe")
    ws3.sheet_properties.tabColor = BLUE

    ws3.merge_cells("A1:F1")
    ws3.cell(
        row=1, column=1,
        value="TRENDY TYGODNIOWE \u2014 porownanie 7d vs 7d"
    ).font = TITLE_FONT
    ws3.merge_cells("A2:F2")
    ws3.cell(
        row=2, column=1,
        value=f"Biezacy: {period_start.strftime(DATE_FMT_PL)} \u2013 {period_end.strftime(DATE_FMT_PL)}  |  "
              f"Poprzedni: {period_prev_start.strftime(DATE_FMT_PL)} \u2013 {period_prev_end.strftime(DATE_FMT_PL)}"
    ).font = SUBTITLE_FONT

    row3 = 4
    trend_headers = [
        "Metryka", "Biezacy 7d", "Poprzedni 7d", "\u0394 zmiana", "\u0394 %", "Trend"
    ]
    for i, h in enumerate(trend_headers):
        ws3.cell(row=row3, column=i + 1, value=h)
    style_header_row(ws3, row3, len(trend_headers))
    trend_header_row = row3

    trend_data = [
        ("Zamowienia", curr["count"], prev["count"], NUM_FORMAT),
        ("Przychod PLN", curr["revenue"], prev["revenue"], PLN_FORMAT),
        ("Srednia wartosc zamowienia", curr["avg_order_value"], prev["avg_order_value"], PLN_FORMAT),
        ("Nowi klienci", curr["new_customers"], prev["new_customers"], NUM_FORMAT),
        ("Powracajacy klienci", curr["returning_customers"], prev["returning_customers"], NUM_FORMAT),
        (
            "Unikalni klienci (total)",
            curr["new_customers"] + curr["returning_customers"],
            prev["new_customers"] + prev["returning_customers"],
            NUM_FORMAT,
        ),
    ]

    for label, val_c, val_p, fmt in trend_data:
        row3 += 1
        ws3.cell(row=row3, column=1, value=label).font = BOLD
        c_curr = ws3.cell(row=row3, column=2, value=val_c)
        c_curr.font = KPI_VALUE_FONT
        c_curr.number_format = fmt
        c_curr.alignment = Alignment(horizontal="center")

        c_prev = ws3.cell(row=row3, column=3, value=val_p)
        c_prev.font = NORMAL
        c_prev.number_format = fmt
        c_prev.alignment = Alignment(horizontal="center")

        abs_change = val_c - val_p
        c_abs = ws3.cell(row=row3, column=4, value=abs_change)
        c_abs.font = NORMAL
        c_abs.number_format = fmt
        c_abs.alignment = Alignment(horizontal="center")

        pct_change = delta_pct(val_c, val_p)
        c_pct = ws3.cell(row=row3, column=5, value=pct_change if pct_change is not None else "n/d")
        if isinstance(pct_change, (int, float)):
            c_pct.number_format = PCT_FORMAT
        c_pct.alignment = Alignment(horizontal="center")

        # Trend arrow + color
        if pct_change is not None:
            if pct_change > 0.05:
                trend_text = "\u2191 wzrost"
                c_pct.fill = GREEN_FILL
                c_pct.font = Font(name="Calibri", bold=True, color="006100")
            elif pct_change < -0.05:
                trend_text = "\u2193 spadek"
                c_pct.fill = RED_FILL
                c_pct.font = Font(name="Calibri", bold=True, color="9C0006")
            else:
                trend_text = "\u2194 stabilnie"
                c_pct.fill = YELLOW_FILL
                c_pct.font = Font(name="Calibri", bold=True, color="9C6500")
        else:
            trend_text = "n/d"

        ws3.cell(row=row3, column=6, value=trend_text).font = NORMAL
        ws3.cell(row=row3, column=6).alignment = Alignment(horizontal="center")

    apply_borders(ws3, trend_header_row, row3, len(trend_headers))

    # Day-by-day comparison
    row3 += 2
    style_section_row(ws3, row3, 6, "PRZYCHOD WG DNIA (biezacy vs poprzedni okres)")
    row3 += 1
    daycmp_headers = ["Dzien tygodnia", "Data (biezacy)", "Przychod (biezacy)", "Data (poprzedni)", "Przychod (poprzedni)", "\u0394 %"]
    for i, h in enumerate(daycmp_headers):
        ws3.cell(row=row3, column=i + 1, value=h)
    style_header_row(ws3, row3, len(daycmp_headers))
    ws3.row_dimensions[row3].height = 30
    daycmp_header_row = row3

    sorted_curr_days = sorted(curr["by_day"].keys())
    sorted_prev_days = sorted(prev["by_day"].keys())
    days_pl = ["Pon", "Wt", "Sr", "Czw", "Pt", "Sob", "Ndz"]

    for idx in range(7):
        row3 += 1
        curr_day = sorted_curr_days[idx] if idx < len(sorted_curr_days) else None
        prev_day = sorted_prev_days[idx] if idx < len(sorted_prev_days) else None

        if curr_day:
            dow = datetime.strptime(curr_day, DATE_FMT).weekday()
            ws3.cell(row=row3, column=1, value=days_pl[dow]).font = NORMAL
        else:
            ws3.cell(row=row3, column=1, value=days_pl[idx] if idx < 7 else "?").font = NORMAL

        ws3.cell(row=row3, column=2, value=curr_day or "\u2014").font = NORMAL
        rev_c = curr["by_day"].get(curr_day, {}).get("revenue", 0) if curr_day else 0
        c_rc = ws3.cell(row=row3, column=3, value=rev_c)
        c_rc.font = NORMAL
        c_rc.number_format = PLN_FORMAT
        c_rc.alignment = Alignment(horizontal="right")

        ws3.cell(row=row3, column=4, value=prev_day or "\u2014").font = NORMAL
        rev_p = prev["by_day"].get(prev_day, {}).get("revenue", 0) if prev_day else 0
        c_rp = ws3.cell(row=row3, column=5, value=rev_p)
        c_rp.font = NORMAL
        c_rp.number_format = PLN_FORMAT
        c_rp.alignment = Alignment(horizontal="right")

        d = delta_pct(rev_c, rev_p)
        c_d = ws3.cell(row=row3, column=6, value=d if d is not None else "n/d")
        if isinstance(d, (int, float)):
            c_d.number_format = PCT_FORMAT
            if d > 0:
                c_d.fill = GREEN_FILL
            elif d < 0:
                c_d.fill = RED_FILL
        c_d.alignment = Alignment(horizontal="center")

    apply_borders(ws3, daycmp_header_row, row3, len(daycmp_headers))

    col_widths_s3 = {1: 28, 2: 18, 3: 20, 4: 18, 5: 20, 6: 14}
    for c, w in col_widths_s3.items():
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.freeze_panes = "A4"

    # ==========================================================
    # SHEET 4: INSTRUKCJA MCP
    # ==========================================================
    print("  Arkusz: Instrukcja MCP...")
    ws4 = wb.create_sheet("Instrukcja MCP")
    ws4.sheet_properties.tabColor = RED

    ws4.merge_cells("A1:E1")
    ws4.cell(
        row=1, column=1,
        value="JAK UZUPELNIC DASHBOARD DANYMI Z MCP"
    ).font = TITLE_FONT
    ws4.merge_cells("A2:E2")
    ws4.cell(
        row=2, column=1,
        value="Instrukcja krok po kroku \u2014 uzyj tych narzedzi MCP w Claude Code, aby wypelnic placeholder-y"
    ).font = SUBTITLE_FONT

    instructions = [
        ("", ""),
        ("KROK 1: GA4 \u2014 Sesje wg kanalu", ""),
        ("Narzedzie:", "mcp__ga4__run_report"),
        ("Parametry:", ""),
        ("  property_id:", "279858535"),
        ("  date_ranges:", f'[{{"start_date": "{period_start.strftime(DATE_FMT)}", "end_date": "{period_end.strftime(DATE_FMT)}"}}]'),
        ("  dimensions:", '["sessionDefaultChannelGroup"]'),
        ("  metrics:", '["sessions", "totalUsers", "conversions", "totalRevenue"]'),
        ("Gdzie wkleic:", 'Arkusz "Dashboard", sekcja RUCH, kolumny B-G'),
        ("Tip:", "Uruchom raport 2x \u2014 raz dla biezacego, raz dla poprzedniego okresu"),
        ("", ""),
        ("KROK 2: Klaviyo \u2014 Kampanie email", ""),
        ("Narzedzie 1:", "mcp__klaviyo__klaviyo_get_campaigns"),
        ("Parametry:", ""),
        ("  channel:", "email"),
        ("Narzedzie 2:", "mcp__klaviyo__klaviyo_get_campaign_report"),
        ("Parametry:", ""),
        ("  campaign_id:", "(z wyniku powyzszego)"),
        ("  conversion_metric_id:", "(uzyj mcp__klaviyo__klaviyo_get_metrics aby znalezc 'Placed Order')"),
        ("  statistics:", '["open-rate", "click-rate", "revenue", "recipient-count", "unique-opens", "unique-clicks"]'),
        ("Gdzie wkleic:", 'Arkusz "Dashboard", sekcja EMAIL'),
        ("", ""),
        ("KROK 3: Google Ads", ""),
        ("Narzedzie:", "mcp__google-ads__run_gaql"),
        ("Parametry:", ""),
        ("  customer_id:", "3393382047"),
        ("  query:", "SELECT campaign.name, metrics.impressions, metrics.clicks, metrics.cost_micros, metrics.conversions, metrics.conversions_value FROM campaign WHERE segments.date DURING LAST_7_DAYS ORDER BY metrics.cost_micros DESC"),
        ("Uwaga:", "cost_micros nalezy podzielic przez 1 000 000 aby uzyskac PLN"),
        ("Gdzie wkleic:", 'Arkusz "Dashboard", sekcja PAID ADS, wiersze Google Ads'),
        ("", ""),
        ("KROK 4: Meta Ads (Facebook + Instagram)", ""),
        ("Narzedzie:", "mcp__meta-ads__get_campaigns"),
        ("Nastepnie:", "mcp__meta-ads__get_insights per campaign"),
        ("Parametry:", ""),
        ("  date_preset:", "last_7d"),
        ("  fields:", "campaign_name, spend, clicks, cpc, actions, action_values"),
        ("Gdzie wkleic:", 'Arkusz "Dashboard", sekcja PAID ADS, wiersze Meta'),
        ("", ""),
        ("KROK 5: GA4 Realtime (opcjonalnie)", ""),
        ("Narzedzie:", "mcp__ga4__run_realtime_report"),
        ("Parametry:", ""),
        ("  property_id:", "279858535"),
        ("  dimensions:", '["unifiedScreenName"]'),
        ("  metrics:", '["activeUsers"]'),
        ("Uzycie:", "Do monitorowania kampanii w czasie rzeczywistym"),
    ]

    row4 = 3
    for label, value in instructions:
        row4 += 1
        if label.startswith("KROK"):
            row4 += 1  # extra spacing before step
            ws4.cell(row=row4, column=1, value=label).font = Font(
                name="Calibri", bold=True, size=12, color=NAVY
            )
            ws4.cell(row=row4, column=2, value=value).font = NORMAL
        elif label.startswith("  "):
            ws4.cell(row=row4, column=1, value=label).font = Font(
                name="Consolas", size=10, color="666666"
            )
            ws4.cell(row=row4, column=2, value=value).font = Font(
                name="Consolas", size=10
            )
        elif label == "Narzedzie:" or label == "Narzedzie 1:" or label == "Narzedzie 2:" or label == "Nastepnie:":
            ws4.cell(row=row4, column=1, value=label).font = BOLD
            ws4.cell(row=row4, column=2, value=value).font = Font(
                name="Consolas", bold=True, size=10, color="0066CC"
            )
        elif label in ("Gdzie wkleic:", "Tip:", "Uwaga:", "Uzycie:"):
            ws4.cell(row=row4, column=1, value=label).font = Font(
                name="Calibri", bold=True, size=10, color=RED
            )
            ws4.cell(row=row4, column=2, value=value).font = NORMAL
        elif label and value:
            ws4.cell(row=row4, column=1, value=label).font = BOLD
            ws4.cell(row=row4, column=2, value=value).font = NORMAL

    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 100

    # ==========================================================
    # SAVE
    # ==========================================================
    wb.save(OUTPUT_FILE)
    print()
    print(f"  Zapisano: {OUTPUT_FILE}")
    print()
    print("=" * 60)
    print("GOTOWE!")
    print(f"Dane Shopify: LIVE ({curr['count']} + {prev['count']} zamowien)")
    print("Dane GA4/Klaviyo/Ads: PLACEHOLDER (uzupelnij wg instrukcji w arkuszu 4)")
    print("=" * 60)


if __name__ == "__main__":
    main()
