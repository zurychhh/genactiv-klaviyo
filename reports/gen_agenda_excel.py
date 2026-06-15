#!/usr/bin/env python3
"""Generate GenActiv call agenda Excel."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Agenda Call — Czerwiec 2026"

# ── Colors ──
PURPLE = "7B2CBF"
DARK_BG = "2D3142"
DARK_ROW = "1E2030"
ALT_ROW = "252840"
HEADER_BG = "3D1F6B"
WHITE = "FFFFFF"
LIGHT_GRAY = "B0B0C0"
GREEN = "22C55E"
YELLOW = "F59E0B"
RED = "EF4444"

# Category colors
CAT_COLORS = {
    "DATA PULL":     "1E3A5F",
    "SEO":           "2D5016",
    "KLAVIYO":       "6B1D1D",
    "BONUS":         "4A3728",
    "PODSUMOWANIE":  "3D3D5C",
}

# ── Styles ──
header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font = Font(name="Calibri", size=10, color="E0E0F0")
data_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="444466"),
    right=Side(style="thin", color="444466"),
    top=Side(style="thin", color="444466"),
    bottom=Side(style="thin", color="444466"),
)

# ── Title row ──
ws.merge_cells("A1:J1")
title_cell = ws["A1"]
title_cell.value = "GENACTIV — WORKING SESSION AGENDA | Czwartek, 12 czerwca 2026 | 60–90 min"
title_cell.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
title_cell.fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:J2")
sub_cell = ws["A2"]
sub_cell.value = "Live execution — Claude Code + MCP | Roadmapa wzrostu +50% YoY"
sub_cell.font = Font(name="Calibri", size=11, color=LIGHT_GRAY, italic=True)
sub_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")

# ── Headers ──
headers = [
    ("Blok / Czas", 14),
    ("Kategoria", 14),
    ("Task", 28),
    ("Opis", 48),
    ("Wykonuje", 18),
    ("Input (kto wpisuje)", 20),
    ("Double Check", 32),
    ("Oczekiwany Outcome", 36),
    ("Status", 14),
    ("Komentarz", 32),
]

for col_idx, (name, width) in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Data ──
tasks = [
    # BLOCK 1: DATA PULL (0:00-0:10)
    {
        "blok": "0:00–0:10",
        "cat": "DATA PULL",
        "task": "GA4 — ruch i konwersje",
        "opis": "Pobranie danych czerwiec vs maj: sesje, CR overall, CR mobile vs desktop, top kanały atrybucji",
        "wykonuje": "Claude Code",
        "input": "Rafał — uruchomienie query",
        "check": "Sprawdzić czy dane za czerwiec mają min. 7 dni. Porównać z dashboardem GA4 w przeglądarce",
        "outcome": "Tabela: sesje, CR, CR mobile, top 5 kanałów — czerwiec vs maj z deltą %",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "DATA PULL",
        "task": "Klaviyo — email performance",
        "opis": "Aktualny rozmiar listy, revenue z flows vs campaigns, open rate, click rate",
        "wykonuje": "Claude Code",
        "input": "Rafał — uruchomienie query",
        "check": "Zweryfikować rozmiar listy w Klaviyo UI. Sprawdzić revenue vs Shopify attribution",
        "outcome": "Metryki: list size, flow revenue 30d, campaign revenue 30d, OR, CTR",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "DATA PULL",
        "task": "SEO Audit — baseline",
        "opis": "Ile issues ogółem, ile high-priority, alt text coverage %, meta title coverage %",
        "wykonuje": "Claude Code",
        "input": "Rafał — uruchomienie audytu",
        "check": "Porównać z ostatnim audytem (457 issues baseline). Liczba powinna być >= 400",
        "outcome": "Baseline: X issues total, Y high, Z% alt text, W% meta titles",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "DATA PULL",
        "task": "Senuto — widoczność domeny",
        "opis": "Aktualna widoczność genactiv.pl, trend 3M, top 10 fraz, pozycje kluczowe",
        "wykonuje": "Claude Code",
        "input": "Rafał — uruchomienie query",
        "check": "Porównać z poprzednim pomiarem. Czy widoczność rośnie/spada?",
        "outcome": "Widoczność: score, trend, top frazy z pozycjami",
        "status": "",
        "comment": "",
    },

    # BLOCK 2: SEO (0:10-0:35)
    {
        "blok": "0:10–0:35",
        "cat": "SEO",
        "task": "Alt texty — bulk fix",
        "opis": "Wygenerowanie alt textow dla 75-100 obrazow (3-4 batche x 25). Format: [Produkt] - [typ] Genactiv",
        "wykonuje": "Claude Code + Human",
        "input": "Rafał — review wygenerowanych alt textów przed apply. Klient — akceptacja formatu",
        "check": "Dry-run → review sample 5 produktów → sprawdzić w Shopify admin czy alt texty się zapisały poprawnie",
        "outcome": "Alt text coverage: 0,3% → 25–30%. Min. 75 obrazów uzupełnionych",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "SEO",
        "task": "Meta titles — bulk fix",
        "opis": "Uzupelnienie 50-59 brakujacych meta titles (2 batche x 25). Format: [Produkt] | Genactiv - [benefit], max 60 znakow",
        "wykonuje": "Claude Code + Human",
        "input": "Rafał — review treści meta titles. Klient — akceptacja tone of voice",
        "check": "Dry-run → review listy → sprawdzić 3 produkty w Shopify admin → Google Rich Results preview",
        "outcome": "Meta title coverage: ~38% → ~80%. Wszystkie top produkty mają meta title",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "SEO",
        "task": "Meta descriptions — bulk fix",
        "opis": "Uzupelnienie 18 brakujacych meta descriptions (1 batch). Format: [Benefit]. [USP]. [CTA]. Od X PLN. 120-155 znakow",
        "wykonuje": "Claude Code + Human",
        "input": "Rafał — review treści. Klient — weryfikacja cen i USP",
        "check": "Dry-run → review → sprawdzić długość (120–155 chars) → podgląd w Google SERP preview",
        "outcome": "0 brakujących meta descriptions. Wszystkie 120–155 znaków z CTA",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "SEO",
        "task": "Re-run SEO Audit — weryfikacja",
        "opis": "Ponowne uruchomienie audytu SEO żeby potwierdzić spadek issues po fixach",
        "wykonuje": "Claude Code",
        "input": "Rafał — uruchomienie",
        "check": "Issues powinny spaść o ~100–130 vs baseline z bloku 1. Jeśli nie — debug",
        "outcome": "Nowy audit: ~300 issues (vs ~415 baseline). Potwierdzenie że fixy się zapisały",
        "status": "",
        "comment": "",
    },

    # BLOCK 3: KLAVIYO (0:35-0:55)
    {
        "blok": "0:35–0:55",
        "cat": "KLAVIYO",
        "task": "Post-Purchase — Email 1 (Thank You)",
        "opis": "Szablon HTML: podziękowanie za zakup + jak stosować produkt + link do FAQ. Trigger: Placed Order + 1 dzień. Brand GenActiv (fonty, kolory, layout)",
        "wykonuje": "Claude Code + Human",
        "input": "Klient — akceptacja copy i layoutu. Rafał — review HTML / responsywność",
        "check": "Podgląd w Klaviyo → test wysyłka na email testowy → sprawdzić mobile rendering → linki działają",
        "outcome": "Szablon utworzony w Klaviyo. Test email wysłany i zweryfikowany",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "KLAVIYO",
        "task": "Post-Purchase — Email 2 (Cross-sell)",
        "opis": "Szablon HTML: cross-sell D+14. Logika: colostrum kupił → pokaż fiberbiom, fiberbiom kupił → pokaż colostrum. Dynamiczne produkty",
        "wykonuje": "Claude Code + Human",
        "input": "Klient — akceptacja par cross-sell. Rafał — review Klaviyo template variables",
        "check": "Podgląd w Klaviyo → test email → sprawdzić czy dynamiczne bloki działają → mobile OK",
        "outcome": "Szablon z dynamic content gotowy. Pary cross-sell skonfigurowane",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "KLAVIYO",
        "task": "Win-Back — Email 1 (reactivation)",
        "opis": "Szablon HTML: Dawno Cie nie widzielismy + bestsellery + 10% kod rabatowy. Trigger: Last Purchase > 90 dni",
        "wykonuje": "Claude Code + Human",
        "input": "Klient — akceptacja kodu rabatowego i % zniżki. Rafał — review copy",
        "check": "Podgląd w Klaviyo → test email → kod rabatowy działa w Shopify → mobile rendering OK",
        "outcome": "Szablon gotowy. Kod rabatowy skonfigurowany w Shopify (jeśli czas pozwoli)",
        "status": "",
        "comment": "",
    },

    # BLOCK 4: BONUS (0:55-1:10)
    {
        "blok": "0:55–1:10",
        "cat": "BONUS",
        "task": "Senuto — competitor gap analysis",
        "opis": "genactiv.pl vs colostrigen.pl: wspólne frazy, unikalne frazy konkurenta, luki do wykorzystania w content",
        "wykonuje": "Claude Code",
        "input": "Rafał — wybór konkurenta do analizy",
        "check": "Czy dane obejmują min. 50 fraz? Czy widać wyraźne luki (frazy gdzie konkurent jest a my nie)?",
        "outcome": "Lista 10–20 keyword gaps do wykorzystania w content strategy (blog, LP)",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "BONUS",
        "task": "Google Ads — przegląd kampanii",
        "opis": "ROAS per kampania, spend, konwersje. Identyfikacja kampanii z niskim ROAS. Rekomendacja realokacji",
        "wykonuje": "Claude Code + Human",
        "input": "Rafał — uruchomienie. Klient — kontekst biznesowy kampanii",
        "check": "Porównać dane MCP z Google Ads UI. Czy ROAS zgadza się z ręcznym obliczeniem?",
        "outcome": "Ranking kampanii po ROAS. Lista 2–3 kampanii do wyłączenia / realokacji budżetu",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "BONUS",
        "task": "Klaviyo — Sunset flow (opcjonalnie)",
        "opis": "Flow 180-day inactive: re-engage email → jeśli brak reakcji → suppress. Higiena listy",
        "wykonuje": "Claude Code + Human",
        "input": "Rafał — review copy. Klient — akceptacja suppression policy",
        "check": "Test email → segment 180d inactive istnieje w Klaviyo → logika suppress poprawna",
        "outcome": "Szablon sunset ready. Segment zdefiniowany (aktywacja po callu)",
        "status": "",
        "comment": "",
    },

    # BLOCK 5: PODSUMOWANIE (1:10-1:20)
    {
        "blok": "1:10–1:20",
        "cat": "PODSUMOWANIE",
        "task": "Review: przed vs po",
        "opis": "Porównanie stanu sprzed sesji z aktualnym: SEO issues, alt text %, meta coverage, flows count",
        "wykonuje": "Claude Code + Human",
        "input": "Obaj — wspólny przegląd",
        "check": "Tabela przed/po zgadza się z faktycznym stanem w Shopify i Klaviyo",
        "outcome": "Raport: delta na każdej metryce. Screenshot/export do dokumentacji",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "PODSUMOWANIE",
        "task": "Aktywacja flows — co zostaje",
        "opis": "Ustalenie co trzeba jeszcze zrobić w Klaviyo UI: ustawić triggery, segmenty, włączyć live",
        "wykonuje": "Human (Klient/Rafał)",
        "input": "Klient — decyzja o aktywacji (natychmiast vs po review)",
        "check": "Checklist: trigger ustawiony, segment poprawny, test email OK, flow = LIVE",
        "outcome": "Lista TODO z ownerami i deadlinem (do końca tygodnia)",
        "status": "",
        "comment": "",
    },
    {
        "blok": "",
        "cat": "PODSUMOWANIE",
        "task": "Plan na lipiec — ustalenie priorytetów",
        "opis": "Na bazie roadmapy: content hub (4 blog posts), A/B test PDP, subskrypcje, VIP flow, schema markup",
        "wykonuje": "Human (obaj)",
        "input": "Klient — priorytety biznesowe. Rafał — feasibility techniczny",
        "check": "Max 5 priorytetów na lipiec. Każdy z ownerem i definicją done",
        "outcome": "Uzgodniona lista top 5 priorytetów lipiec z ownerami",
        "status": "",
        "comment": "",
    },
]

# ── Write data rows ──
for row_idx, t in enumerate(tasks, 4):
    is_odd = (row_idx % 2 == 0)
    row_fill = PatternFill(start_color=DARK_ROW if is_odd else ALT_ROW,
                           end_color=DARK_ROW if is_odd else ALT_ROW,
                           fill_type="solid")

    cat_color = CAT_COLORS.get(t["cat"], DARK_ROW)
    cat_fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type="solid")

    values = [
        t["blok"], t["cat"], t["task"], t["opis"],
        t["wykonuje"], t["input"], t["check"], t["outcome"],
        t["status"], t["comment"],
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_align if col_idx >= 3 else center_align
        cell.border = thin_border

        if col_idx == 2:
            cell.fill = cat_fill
            cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            cell.alignment = center_align
        elif col_idx == 9:
            cell.fill = row_fill
            cell.alignment = center_align
        else:
            cell.fill = row_fill

    # Bold task name
    ws.cell(row=row_idx, column=3).font = Font(name="Calibri", bold=True, size=10, color=WHITE)

# ── Status dropdown ──
status_validation = DataValidation(
    type="list",
    formula1='"Done,In Progress,On Hold,Skipped"',
    allow_blank=True,
)
status_validation.error = "Wybierz status z listy"
status_validation.errorTitle = "Status"
ws.add_data_validation(status_validation)

last_row = 3 + len(tasks)
status_validation.add(f"I4:I{last_row}")

# ── Freeze panes ──
ws.freeze_panes = "A4"

# ── Row height ──
ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 24
ws.row_dimensions[3].height = 32
for r in range(4, last_row + 1):
    ws.row_dimensions[r].height = 56

# ── Sheet tab color ──
ws.sheet_properties.tabColor = PURPLE

# ── Save ──
output_path = "/Users/user/Downloads/Genactiv/agenda-call-czerwiec-2026.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
