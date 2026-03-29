#!/usr/bin/env python3
"""
Generuje Excel z porównaniem źródeł ruchu: GA4 vs Shopify vs Klaviyo
Okres: 20-26.02.2026
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
HEADER_FILL = PatternFill(start_color="1A3B5D", end_color="1A3B5D", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
GAP_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # yellow
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1A3B5D")
BOLD = Font(name="Calibri", bold=True, size=10)
NORMAL = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
PLN_FORMAT = '#,##0'
PCT_FORMAT = '0%'

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_subheader_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_borders(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

def auto_width(ws, max_col, min_width=10, max_width=35):
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len


# ============================================================
# SHEET 1: POROWNANIE ZRODEL (cross-reference)
# ============================================================
ws1 = wb.active
ws1.title = "1. Porównanie źródeł"
ws1.sheet_properties.tabColor = "1A3B5D"

ws1.cell(row=1, column=1, value="WERYFIKACJA ŹRÓDEŁ RUCHU — GenActiv.pl").font = TITLE_FONT
ws1.cell(row=2, column=1, value="Okres: 20-26.02.2026 | Wygenerowano: " + datetime.now().strftime("%Y-%m-%d %H:%M")).font = NORMAL

headers = [
    "Kanał / Źródło",
    "GA4\nSesje", "GA4\nUżytkownicy", "GA4\nKonwersje", "GA4\nPrzychód PLN",
    "Shopify Journey\nZamówienia\n(last visit)", "Shopify Journey\nZ UTM",
    "Shopify Orders\n(MCP)\nZamówienia", "Shopify Orders\nPrzychód PLN",
    "Δ Konwersje\nGA4 vs Shopify", "Status\nweryfikacji", "Uwagi"
]
row = 4
for i, h in enumerate(headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, len(headers))
ws1.row_dimensions[row].height = 45

# Data rows - manually compiled from all 3 sources
data = [
    # [kanał, ga4_sesje, ga4_users, ga4_conv, ga4_rev, shop_journey_orders, shop_journey_utm, shop_mcp_orders, shop_mcp_rev, delta, status, uwagi]
    ["Google Organic (SEO)", 2488, 2097, 40, 11902, 113, 0, 129, 37748, "GA4: 40 vs SH: 113-129", "ROZBIEŻNOŚĆ", "GA4 traci ~65% konwersji. Consent Mode + Przelewy24 redirect przerywa sesję."],
    ["Google Ads (CPC / PMax)", 2617, 1930, 65, 16870, "—", "—", "—", "—", "—", "OK (gclid)", "Auto-tagging działa. Cross-network w GA4 obejmuje PMax."],
    ["Direct", 667, 557, 20, 4221, 77, 1, 55, 16146, "GA4: 20 vs SH: 55-77", "ROZBIEŻNOŚĆ", "Shopify liczy więcej direct. GA4 traci przez consent."],
    ["Instagram (paid)", 234, 210, 6, 1143, 6, 6, 21, 4847, "GA4: 6 vs SH: 6-21", "CZĘŚCIOWO OK", "Paid z UTM się zgadza. Organic IG wchodzi jako referral."],
    ["Instagram (organic)", 309, 283, 13, 3060, 12, 0, "—", "—", "—", "GAP: brak UTM", "309 sesji z l.instagram.com/referral. 12 zamówień BEZ UTM w Shopify."],
    ["Facebook (paid)", 64, 61, 1, 169, 3, 3, 8, 1904, "GA4: 1 vs SH: 3-8", "ROZBIEŻNOŚĆ", "GA4 traci konwersje. Facebook + case mismatch (F/f)."],
    ["Facebook (organic)", 53, 44, 0, 0, 5, 0, "—", "—", "—", "GAP: brak UTM", "m.facebook.com/referral w GA4. Brak UTM na postach organicznych."],
    ["Email: newsletter/*", 88, 69, 4, 1132, 16, 16, 3, 1421, "GA4: 4 vs SH: 3-16", "CZĘŚCIOWO OK", "utm_source=newsletter. Case mismatch: JESIENGENACTIV vs JesienGenactiv."],
    ["Email: Klaviyo/*", 11, 7, 1, 0, 1, 1, 5, 1751, "GA4: 1 vs SH: 1-5", "ROZBIEŻNOŚĆ", "utm_source=Klaviyo. GA4 traci konwersje email."],
    ["Email: Gmail App", "—", "—", "—", "—", 4, 0, "—", "—", "—", "GAP: brak UTM", "android-app://com.google.android.gm/ stripuje UTMy."],
    ["Bing Ads (CPC)", 270, 207, 6, 2365, 3, 2, 2, 982, "GA4: 6 vs SH: 2-3", "OK", "UTMy prawidłowe (Sembot). GA4 liczy assisted conversions."],
    ["Bing Organic", 13, 12, 0, 0, 1, 0, 1, 106, "—", "OK", "Mały wolumen."],
    ["TikTok (paid)", 356, 356, 0, 0, 0, 0, 0, 0, "0 vs 0", "OK (brak konwersji)", "Duży ruch, zero konwersji. UTMy obecne."],
    ["TikTok (paid social)", 77, 71, 0, 0, 0, 0, 0, 0, "0 vs 0", "OK", "junior_winter_2026 campaign."],
    ["YouTube (organic)", 453, 388, 0, 0, 1, 1, 1, 340, "GA4: 0 vs SH: 1", "MINOR", "Jedno zamówienie, GA4 nie przypisało."],
    ["Unassigned (GA4)", 1491, 1216, 21, 6202, "—", "—", "—", "—", "—", "PROBLEM", "15% sesji bez atrybucji. Consent Mode v2 blokuje analytics_storage."],
    ["(not set) / 3528", 301, "—", 11, 3021, "—", "—", "—", "—", "—", "PROBLEM", "Kampania '3528' - niezidentyfikowana. Sprawdzić źródło."],
    ["Inne (teads, polsat, rmf, qrkod, librus, ulotka)", 83, "—", 0, 0, 3, 0, "—", "—", "—", "OK", "Małe wolumeny, offline sources."],
]

for i, d in enumerate(data):
    r = row + 1 + i
    for j, val in enumerate(d):
        cell = ws1.cell(row=r, column=j+1, value=val)
        cell.font = NORMAL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if j in [1, 2, 3, 5, 6, 7]:  # numeric
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if j in [4, 8]:  # PLN
            if isinstance(val, (int, float)):
                cell.number_format = PLN_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Color status column
    status_cell = ws1.cell(row=r, column=11)
    status = str(status_cell.value)
    if "ROZBIEŻNOŚĆ" in status or "PROBLEM" in status:
        status_cell.fill = GAP_FILL
        status_cell.font = Font(name="Calibri", bold=True, size=10, color="9C0006")
    elif "GAP" in status:
        status_cell.fill = WARN_FILL
        status_cell.font = Font(name="Calibri", bold=True, size=10, color="9C6500")
    elif "OK" in status:
        status_cell.fill = OK_FILL
        status_cell.font = Font(name="Calibri", bold=True, size=10, color="006100")
    elif "CZĘŚCIOWO" in status:
        status_cell.fill = WARN_FILL
        status_cell.font = Font(name="Calibri", bold=True, size=10, color="9C6500")

# Totals row
r_total = row + 1 + len(data)
ws1.cell(row=r_total, column=1, value="SUMA (GA4 total)").font = BOLD
ws1.cell(row=r_total, column=2, value=9683).font = BOLD
ws1.cell(row=r_total, column=3, value="—").font = BOLD
ws1.cell(row=r_total, column=4, value=181).font = BOLD
ws1.cell(row=r_total, column=5, value=47749).font = BOLD
ws1.cell(row=r_total, column=5).number_format = PLN_FORMAT
ws1.cell(row=r_total, column=6, value=264).font = BOLD
ws1.cell(row=r_total, column=8, value=250).font = BOLD
ws1.cell(row=r_total, column=9, value=72082).font = BOLD
ws1.cell(row=r_total, column=9).number_format = PLN_FORMAT

apply_borders(ws1, row, r_total, len(headers))
auto_width(ws1, len(headers))
ws1.freeze_panes = "B5"


# ============================================================
# SHEET 2: AUDYT UTM
# ============================================================
ws2 = wb.create_sheet("2. Audyt UTM")
ws2.sheet_properties.tabColor = "EF3340"

ws2.cell(row=1, column=1, value="AUDYT UTM — co wysyłają kampanie vs co rejestrują narzędzia").font = TITLE_FONT
ws2.cell(row=2, column=1, value="Okres: 20-26.02.2026").font = NORMAL

headers2 = [
    "Platforma\nkampanii", "utm_source", "utm_medium", "utm_campaign",
    "utm_content", "utm_term",
    "GA4\nSesje", "GA4\nKonwersje",
    "Shopify Journey\nZamówienia",
    "Spójność\nsource/medium", "Status", "Problem / Uwagi"
]
r2 = 4
for i, h in enumerate(headers2, 1):
    ws2.cell(row=r2, column=i, value=h)
style_header_row(ws2, r2, len(headers2))
ws2.row_dimensions[r2].height = 45

utm_data = [
    # [platforma, source, medium, campaign, content, term, ga4_sess, ga4_conv, shop_orders, spójność, status, problem]
    ["Klaviyo (nowe)", "Klaviyo", "email", "ZimazColostrum", "(brak)", "(brak)", 5, 0, 1, "OK", "WARN", "Brak utm_content i utm_term. Source ≠ 'newsletter'."],
    ["Klaviyo (nowe)", "Klaviyo", "email", "ColostrumA2", "(brak)", "Colostrum A2...", 4, 1, "—", "OK", "WARN", "Brak utm_content. Term OK."],
    ["Klaviyo (nowe)", "Klaviyo", "email", "BLACKFRIDAY", "(brak)", "Colostrum", 2, 0, "—", "OK", "OK", "Stara kampania, nadal generuje kliknięcia."],
    ["Klaviyo (stare)", "newsletter", "email", "JESIENGENACTIV", "(brak)", "Colostrum", 66, 3, 12, "INNY SOURCE!", "GAP", "utm_source=newsletter zamiast Klaviyo. Rozbija raportowanie."],
    ["Klaviyo (stare)", "newsletter", "email", "JesienGenactiv", "(brak)", "Colostrum", 22, 1, 4, "CASE MISMATCH!", "GAP", "Ta sama kampania co wyżej ale inny case! GA4 liczy osobno."],
    ["Klaviyo (stare)", "newsletter", "(brak!)", "(brak)", "(brak)", "(brak)", 4, 0, "—", "BRAK MEDIUM!", "GAP", "4 sesje z utm_source=newsletter ale BEZ utm_medium."],
    ["Klaviyo (luty)", "?", "?", "?", "?", "?", 0, 0, "—", "NIEWIDOCZNE!", "GAP KRYTYCZNY", "Kampanie A2, Junior, Follow-up z lutego 2026 - 0 sesji w GA4!"],
    ["Meta Ads", "instagram", "paidsocial", "junior_winter_2026", "animated_ad", "sales", 234, 6, 5, "OK", "OK", "UTMy kompletne i spójne."],
    ["Meta Ads", "facebook", "paidsocial", "junior_winter_2026", "animated_ad", "sales", 64, 1, 3, "OK", "OK", "UTMy kompletne."],
    ["Meta Ads", "Facebook", "paidsocial", "junior_winter_2026", "?", "?", 9, 0, "—", "CASE MISMATCH!", "GAP", "Uppercase 'F' w Facebook. Inna kampania/ad set?"],
    ["Meta Ads", "Instagram", "cpc", "ZimazColostrum", "?", "?", 9, 0, "—", "INNY MEDIUM!", "GAP", "Medium='cpc' zamiast 'paidsocial'. Niespójność."],
    ["Meta Ads", "Meta", "cpc", "The BEST GENACTIV...", "?", "?", 4, 0, "—", "INNY SOURCE!", "GAP", "Source='Meta' zamiast instagram/facebook. Niespójne."],
    ["Instagram (organic)", "Instagram", "social", "ColostrumA2", "Stories", "Colostrum A2...", 59, 0, 1, "OK", "OK", "Linki ze Stories z UTM - prawidłowe."],
    ["Instagram (organic)", "Instagram", "Bio", "Stan jelit", "(brak)", "(brak)", 7, 0, "—", "NIEST. MEDIUM", "WARN", "Medium='Bio' zamiast 'social'. Niestandardowe."],
    ["Instagram (referral)", "l.instagram.com", "referral", "(referral)", "(brak)", "(brak)", 309, 13, 12, "BRAK UTM!", "GAP", "309 sesji bez UTM! Linki z postów/bio bez parametrów."],
    ["Facebook (referral)", "m.facebook.com", "referral", "(referral)", "(brak)", "(brak)", 53, 0, 5, "BRAK UTM!", "GAP", "Posty organiczne bez UTM."],
    ["Google Ads", "google", "cpc", "Performance Max_*", "(auto)", "(auto)", 2062, 63, "—", "OK (gclid)", "OK", "Auto-tagging. Nie wymaga UTM."],
    ["Google Ads", "google", "cpc", "ProduktowaTest", "(auto)", "(auto)", 263, 2, "—", "OK (gclid)", "OK", ""],
    ["Google Ads", "google", "cpc", "Remarketing_Test", "(auto)", "(auto)", 120, 1, "—", "OK (gclid)", "OK", ""],
    ["Google Ads", "google", "cpc", "(not set)", "(auto)", "(auto)", 7, 0, "—", "BRAK KAMPANII", "WARN", "7 sesji bez nazwy kampanii. Sprawdzić tracking template."],
    ["Bing Ads (Sembot)", "bing", "cpc", "[Sembot][Search] Pure Brand", "Pure Brand", "colostrum genactiv", 49, 4, 2, "OK", "OK", "UTMy kompletne z content i term."],
    ["Bing Ads (Sembot)", "bing", "cpc", "[Sembot][Feed] Catch All", "Produkty", "genactiv.pl", 31, 2, 2, "OK", "OK", ""],
    ["Bing Ads (Sembot)", "bing", "cpc", "[Sembot][Audience] Ogólna", "?", "?", 126, 0, "—", "OK", "OK", ""],
    ["TikTok Ads", "tiktok", "cpc", "ZimazColostrum", "?", "?", 356, 0, 0, "OK", "OK", "Duży ruch, 0 konwersji."],
    ["TikTok Ads", "tiktok", "paidsocial", "junior_winter_2026", "?", "?", 77, 0, 0, "INNY MEDIUM", "WARN", "Medium='paidsocial' vs 'cpc' w innej kampanii TikTok."],
    ["Gmail App (email)", "android-app://...", "(brak)", "(brak)", "(brak)", "(brak)", "—", "—", 4, "UTM STRIPPED", "GAP", "Gmail app na Android usuwa UTMy z linków email."],
    ["Google (product sync)", "google", "product_sync", "sag_organic", "sag_organic", "(brak)", 4, 0, 3, "OK", "OK", "Shopify Google Channel. Nie Google Ads."],
    ["Offline (QR, ulotka)", "qrkod/ulotka/*", "probka/kodqr/*", "różne", "—", "—", 27, 0, "—", "OK", "OK", "Małe wolumeny, UTMy obecne."],
]

for i, d in enumerate(utm_data):
    r = r2 + 1 + i
    for j, val in enumerate(d):
        cell = ws2.cell(row=r, column=j+1, value=val)
        cell.font = NORMAL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if j in [6, 7, 8]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    status_cell = ws2.cell(row=r, column=11)
    status = str(status_cell.value)
    if "GAP" in status:
        status_cell.fill = GAP_FILL
        status_cell.font = Font(name="Calibri", bold=True, size=10, color="9C0006")
    elif "WARN" in status:
        status_cell.fill = WARN_FILL
        status_cell.font = Font(name="Calibri", bold=True, size=10, color="9C6500")
    elif "OK" in status:
        status_cell.fill = OK_FILL
        status_cell.font = Font(name="Calibri", bold=True, size=10, color="006100")

    # Also color the "spójność" column
    sp_cell = ws2.cell(row=r, column=10)
    sp = str(sp_cell.value)
    if "BRAK" in sp or "MISMATCH" in sp or "INNY" in sp or "NIEWIDOCZNE" in sp or "STRIPPED" in sp:
        sp_cell.fill = GAP_FILL
    elif "NIEST" in sp:
        sp_cell.fill = WARN_FILL
    elif "OK" in sp:
        sp_cell.fill = OK_FILL

apply_borders(ws2, r2, r2 + len(utm_data), len(headers2))
auto_width(ws2, len(headers2))
ws2.freeze_panes = "B5"


# ============================================================
# SHEET 3: KLAVIYO KAMPANIE - konfiguracja UTM
# ============================================================
ws3 = wb.create_sheet("3. Klaviyo kampanie")
ws3.sheet_properties.tabColor = "27AE60"

ws3.cell(row=1, column=1, value="KLAVIYO — Konfiguracja kampanii email (styczeń-luty 2026)").font = TITLE_FONT

headers3 = [
    "Nazwa kampanii", "Data wysyłki", "Status",
    "add_tracking_params", "custom_tracking_params",
    "Widoczna w GA4?", "utm_source w GA4", "utm_campaign w GA4",
    "Sesje GA4", "Status UTM", "Akcja wymagana"
]
r3 = 3
for i, h in enumerate(headers3, 1):
    ws3.cell(row=r3, column=i, value=h)
style_header_row(ws3, r3, len(headers3))
ws3.row_dimensions[r3].height = 35

klaviyo_data = [
    ["[Follow-up] Kampania – 12 lut 2026", "2026-02-14", "Sent", "TAK", "[] (puste!)", "NIE!", "—", "—", 0, "GAP KRYTYCZNY", "Sprawdzić linki w szablonie - brak UTM?"],
    ["Kampania – 12 lut 2026, 13:23", "2026-02-12", "Sent", "TAK", "[] (puste!)", "NIE!", "—", "—", 0, "GAP KRYTYCZNY", "Sprawdzić linki w szablonie - brak UTM?"],
    ["[Follow-up] A2", "2026-02-07", "Sent", "TAK", "[] (puste!)", "NIE!", "—", "—", 0, "GAP KRYTYCZNY", "Sprawdzić linki w szablonie - brak UTM?"],
    ["A2", "2026-02-05", "Sent", "TAK", "[] (puste!)", "TAK (prawdop.)", "Klaviyo", "ColostrumA2", 4, "OK (ale brak content)", "Dodać utm_content"],
    ["JUNIOR", "2026-01-29", "Sent", "TAK", "[] (puste!)", "?", "?", "?", "?", "DO WERYFIKACJI", "Sprawdzić w GA4 za okres 29.01"],
    ["[Follow-up] ZOOOGGGIEES", "2026-01-24", "Sent", "TAK", "[] (puste!)", "?", "?", "?", "?", "DO WERYFIKACJI", "Sprawdzić w GA4 za okres 24.01"],
    ["ZOOOGGGIEES", "2026-01-22", "Sent", "TAK", "[] (puste!)", "?", "?", "?", "?", "DO WERYFIKACJI", "Sprawdzić w GA4 za okres 22.01"],
    ["Colostrum w płynie", "2026-01-15", "Sent", "TAK", "[] (puste!)", "?", "?", "?", "?", "DO WERYFIKACJI", "Sprawdzić w GA4 za okres 15.01"],
    ["Nowość_MieszankaCIMK", "2026-01-08", "Sent", "TAK", "[] (puste!)", "?", "?", "?", "?", "DO WERYFIKACJI", "Sprawdzić w GA4 za okres 8.01"],
]

for i, d in enumerate(klaviyo_data):
    r = r3 + 1 + i
    for j, val in enumerate(d):
        cell = ws3.cell(row=r, column=j+1, value=val)
        cell.font = NORMAL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    status_cell = ws3.cell(row=r, column=10)
    status = str(status_cell.value)
    if "GAP" in status:
        status_cell.fill = GAP_FILL
        status_cell.font = Font(name="Calibri", bold=True, size=10, color="9C0006")
    elif "WERYFIKACJI" in status:
        status_cell.fill = WARN_FILL
        status_cell.font = Font(name="Calibri", bold=True, size=10, color="9C6500")
    elif "OK" in status:
        status_cell.fill = OK_FILL

apply_borders(ws3, r3, r3 + len(klaviyo_data), len(headers3))
auto_width(ws3, len(headers3))
ws3.freeze_panes = "B4"


# ============================================================
# SHEET 4: GA4 RAW DATA
# ============================================================
ws4 = wb.create_sheet("4. GA4 szczegółowe")
ws4.sheet_properties.tabColor = "4472C4"

ws4.cell(row=1, column=1, value="GA4 — Sesje wg source/medium/campaign (20-26.02.2026)").font = TITLE_FONT

headers4 = ["sessionSource", "sessionMedium", "sessionCampaignName", "Sesje", "Konwersje", "Przychód PLN"]
r4 = 3
for i, h in enumerate(headers4, 1):
    ws4.cell(row=r4, column=i, value=h)
style_header_row(ws4, r4, len(headers4))

ga4_raw = [
    ["google", "organic", "(organic)", 2470, 40, 11902],
    ["google", "cpc", "Performance Max_GENACTIV_COLOSTRUM", 2062, 63, 16341],
    ["(not set)", "(not set)", "(not set)", 1161, 10, 3181],
    ["(direct)", "(none)", "(direct)", 649, 20, 4221],
    ["youtube.com", "referral", "(referral)", 430, 0, 0],
    ["(data not available)", "(data not available)", "(cross-network)", 367, 1, 189],
    ["tiktok", "cpc", "ZimazColostrum", 356, 0, 0],
    ["l.instagram.com", "referral", "(referral)", 309, 13, 3060],
    ["(not set)", "(not set)", "3528", 301, 11, 3021],
    ["google", "cpc", "ProduktowaTest", 263, 2, 174],
    ["instagram", "paidsocial", "junior_winter_2026", 234, 6, 1143],
    ["bing", "cpc", "[Sembot][Audience] Ogólna", 126, 0, 0],
    ["google", "cpc", "Remarketing_Test", 120, 1, 340],
    ["tiktok", "paidsocial", "junior_winter_2026", 77, 0, 0],
    ["google", "cpc", "UGC_Junior", 73, 0, 0],
    ["newsletter", "email", "JESIENGENACTIV", 66, 3, 901],
    ["facebook", "paidsocial", "junior_winter_2026", 64, 1, 169],
    ["Instagram", "social", "ColostrumA2", 59, 0, 0],
    ["m.facebook.com", "referral", "(referral)", 53, 0, 0],
    ["bing", "cpc", "[Sembot][Search] Pure Brand", 49, 4, 1871],
    ["admin.shopify.com", "referral", "(referral)", 39, 1, 189],
    ["bing", "cpc", "[Sembot][Feed] Bestsellery", 37, 0, 0],
    ["bing", "cpc", "[Sembot][Feed] Catch All", 31, 2, 494],
    ["m.youtube.com", "referral", "(referral)", 23, 0, 0],
    ["newsletter", "email", "JesienGenactiv", 22, 1, 231],
    ["(direct)", "(none)", "3528", 18, 0, 0],
    ["bing", "cpc", "[Sembot][Search] Colostrum - ogólne", 14, 0, 0],
    ["bing", "organic", "(organic)", 13, 0, 0],
    ["bing", "cpc", "[Sembot][Search] Produkty", 12, 0, 0],
    ["rmf", "fixed", "2026", 12, 0, 0],
    ["qrkod", "probka", "probki", 11, 0, 0],
    ["teads", "CPCV", "2026Q1_red", 11, 0, 0],
    ["Facebook", "paidsocial", "junior_winter_2026", 9, 0, 0],
    ["Instagram", "cpc", "ZimazColostrum", 9, 0, 0],
    ["LIBRUS", "Banner", "ZIMAZCOLOSTRUM", 9, 0, 0],
    ["facebook.com", "referral", "(referral)", 8, 0, 0],
    ["polsat_vod", "CPM", "2026Q1_red", 8, 0, 0],
    ["Instagram", "Bio", "Stan jelit", 7, 0, 0],
    ["Ulotka_produktowa", "Colostrum", "opakowania_ulotka", 7, 0, 0],
    ["google", "cpc", "(not set)", 7, 0, 0],
    ["l.facebook.com", "referral", "(referral)", 7, 0, 0],
    ["facebook", "referral", "(referral)", 6, 0, 0],
    ["Klaviyo", "email", "ZimazColostrum", 5, 0, 0],
    ["polsat_vod", "CPM", "2026Q1_junior", 5, 0, 0],
    ["ulotka", "kodqr", "ColostrumA2", 5, 0, 0],
    ["Klaviyo", "email", "ColostrumA2", 4, 1, 0],
    ["Meta", "cpc", "The BEST GENACTIV®Colostrum", 4, 0, 0],
    ["QR kod", "probka", "Dobrypoczatek", 4, 0, 0],
    ["chatgpt.com", "(not set)", "(not set)", 4, 0, 0],
    ["google", "product_sync", "sag_organic", 4, 0, 0],
]

for i, d in enumerate(ga4_raw):
    r = r4 + 1 + i
    for j, val in enumerate(d):
        cell = ws4.cell(row=r, column=j+1, value=val)
        cell.font = NORMAL
        if j >= 3:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if j == 5 and isinstance(val, (int, float)):
            cell.number_format = PLN_FORMAT

apply_borders(ws4, r4, r4 + len(ga4_raw), len(headers4))
auto_width(ws4, len(headers4))
ws4.freeze_panes = "A4"


# ============================================================
# SHEET 5: SHOPIFY JOURNEY RAW
# ============================================================
ws5 = wb.create_sheet("5. Shopify Journey")
ws5.sheet_properties.tabColor = "95BF47"

ws5.cell(row=1, column=1, value="SHOPIFY — Customer Journey Last Visit (20-26.02.2026, 264 zamówienia)").font = TITLE_FONT

headers5 = ["Źródło (Shopify)", "sourceType", "Zamówienia (last visit)", "Z UTM (kompletnym)", "Bez UTM", "% bez UTM"]
r5 = 3
for i, h in enumerate(headers5, 1):
    ws5.cell(row=r5, column=i, value=h)
style_header_row(ws5, r5, len(headers5))

journey_data = [
    ["Google", "SEO", 113, 0, 113, 1.0],
    ["direct", "?", 77, 1, 76, 0.987],
    ["Instagram", "?", 18, 6, 12, 0.667],
    ["email", "?", 10, 10, 0, 0.0],
    ["Facebook", "?", 8, 3, 5, 0.625],
    ["android-app://com.google.android.gm/", "?", 4, 0, 4, 1.0],
    ["https://zasobygwp.pl/", "?", 3, 0, 3, 1.0],
    ["Google", "?", 3, 3, 0, 0.0],
    ["Bing", "?", 3, 2, 1, 0.333],
    ["https://account.genactiv.pl/", "?", 2, 0, 2, 1.0],
    ["Yahoo", "?", 1, 0, 1, 1.0],
    ["https://youtube.com/", "?", 1, 1, 0, 0.0],
    ["SUMA", "", 264, 34, 209, 0.792],
]

for i, d in enumerate(journey_data):
    r = r5 + 1 + i
    for j, val in enumerate(d):
        cell = ws5.cell(row=r, column=j+1, value=val)
        cell.font = BOLD if i == len(journey_data) - 1 else NORMAL
        if j >= 2:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if j == 5 and isinstance(val, float):
            cell.number_format = PCT_FORMAT
            if val > 0.5:
                cell.fill = GAP_FILL
            elif val > 0:
                cell.fill = WARN_FILL
            else:
                cell.fill = OK_FILL

# Add UTM combos sub-table
r_sub = r5 + len(journey_data) + 3
ws5.cell(row=r_sub, column=1, value="UTM COMBOS — Last Visit (tylko zamówienia z UTM)").font = TITLE_FONT

headers5b = ["utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term", "Zamówienia"]
r_sub += 1
for i, h in enumerate(headers5b, 1):
    ws5.cell(row=r_sub, column=i, value=h)
style_header_row(ws5, r_sub, len(headers5b))

utm_combos = [
    ["newsletter", "email", "JESIENGENACTIV", "(empty)", "Colostrum", 12],
    ["instagram", "paidsocial", "junior_winter_2026", "animated_ad", "sales", 5],
    ["newsletter", "email", "JesienGenactiv", "(empty)", "Colostrum", 4],
    ["facebook", "paidsocial", "junior_winter_2026", "animated_ad", "sales", 3],
    ["google", "product_sync", "sag_organic", "sag_organic", "(empty)", 3],
    ["bing", "cpc", "[Sembot][Feed] Catch All", "Produkty", "genactiv.pl", 2],
    ["bing", "cpc", "[Sembot][Search] Pure Brand", "Pure Brand", "colostrum genactiv", 2],
    ["Instagram", "social", "ColostrumA2", "Stories", "Colostrum A2...", 1],
    ["Klaviyo", "email", "ZimazColostrum", "(empty)", "(empty)", 1],
    ["Youtube", "Social", "ZimazColostrum", "shorts", "Colostrum...", 1],
]

for i, d in enumerate(utm_combos):
    r = r_sub + 1 + i
    for j, val in enumerate(d):
        cell = ws5.cell(row=r, column=j+1, value=val)
        cell.font = NORMAL
        if val == "(empty)":
            cell.fill = WARN_FILL

apply_borders(ws5, r5, r5 + len(journey_data), len(headers5))
apply_borders(ws5, r_sub, r_sub + len(utm_combos), len(headers5b))
auto_width(ws5, len(headers5))
ws5.freeze_panes = "A4"


# ============================================================
# SHEET 6: CHECKLIST AKCJI
# ============================================================
ws6 = wb.create_sheet("6. Checklist akcji")
ws6.sheet_properties.tabColor = "EF3340"

ws6.cell(row=1, column=1, value="CHECKLIST — Akcje do wykonania").font = TITLE_FONT

headers6 = ["#", "Priorytet", "Akcja", "Dotyczy", "Oczekiwany efekt", "Status"]
r6 = 3
for i, h in enumerate(headers6, 1):
    ws6.cell(row=r6, column=i, value=h)
style_header_row(ws6, r6, len(headers6))

actions = [
    [1, "KRYTYCZNY", "Sprawdzić szablony HTML lutowych kampanii Klaviyo (A2, Junior, Follow-up) — czy linki mają UTM parametry", "Klaviyo", "Zidentyfikować dlaczego 0 sesji w GA4 z lutowych kampanii", "DO ZROBIENIA"],
    [2, "WYSOKI", "Ustandaryzować utm_source na 'klaviyo' (lowercase) we WSZYSTKICH kampaniach email", "Klaviyo", "Jedno źródło 'klaviyo/email' zamiast rozbitego 'newsletter' + 'Klaviyo'", "DO ZROBIENIA"],
    [3, "WYSOKI", "Dodać go.przelewy24.pl do Unwanted Referrals w GA4", "GA4 Admin", "Zmniejszy utratę atrybucji po powrocie z bramki płatności", "DO ZROBIENIA"],
    [4, "WYSOKI", "Dodać UTMy do linku w bio Instagram (utm_source=instagram&utm_medium=social&utm_campaign=bio_link)", "Instagram", "Odzyska 309+ sesji z l.instagram.com/referral", "DO ZROBIENIA"],
    [5, "ŚREDNI", "Ustandaryzować case w Meta Ads URL parameters: lowercase source (instagram/facebook)", "Meta Ads Manager", "Koniec fragmentacji: 'Facebook' vs 'facebook', 'Instagram' vs 'instagram'", "DO ZROBIENIA"],
    [6, "ŚREDNI", "Ustandaryzować medium w Meta Ads: 'paidsocial' dla paid, 'social' dla organic", "Meta Ads + IG linki", "Koniec mieszania 'cpc' / 'paidsocial' / 'social' / 'Bio'", "DO ZROBIENIA"],
    [7, "ŚREDNI", "Dodać utm_content do kampanii Klaviyo (np. wariant szablonu, CTA type)", "Klaviyo", "Możliwość A/B testowania szablonów email", "DO ZROBIENIA"],
    [8, "NISKI", "Zbadać kampanię '3528' w GA4 (301 sesji, 11 konwersji, 3021 PLN)", "GA4 + ?", "Zidentyfikować źródło tych sesji", "DO ZROBIENIA"],
    [9, "NISKI", "Dodać UTMy do linków na zasobygwp.pl", "Partner", "Odzyska 3 zamówienia z referral bez UTM", "DO ZROBIENIA"],
    [10, "NISKI", "Zweryfikować consent banner Pandectes — 15% sesji 'Unassigned'", "Pandectes / GTM", "Zmniejszyć bucket Unassigned w GA4", "DO ZROBIENIA"],
]

for i, d in enumerate(actions):
    r = r6 + 1 + i
    for j, val in enumerate(d):
        cell = ws6.cell(row=r, column=j+1, value=val)
        cell.font = NORMAL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    prio_cell = ws6.cell(row=r, column=2)
    prio = str(prio_cell.value)
    if "KRYTYCZNY" in prio:
        prio_cell.fill = GAP_FILL
        prio_cell.font = Font(name="Calibri", bold=True, size=10, color="9C0006")
    elif "WYSOKI" in prio:
        prio_cell.fill = WARN_FILL
        prio_cell.font = Font(name="Calibri", bold=True, size=10, color="9C6500")
    elif "ŚREDNI" in prio:
        prio_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    elif "NISKI" in prio:
        prio_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    status_cell = ws6.cell(row=r, column=6)
    status_cell.fill = WARN_FILL

apply_borders(ws6, r6, r6 + len(actions), len(headers6))
auto_width(ws6, len(headers6))
ws6.column_dimensions["C"].width = 70
ws6.column_dimensions["E"].width = 50
ws6.freeze_panes = "A4"


# ============================================================
# SAVE
# ============================================================
filepath = "/Users/user/projects/genactiv-klaviyo/reports/weryfikacja_zrodel_ruchu_20-26_02_2026.xlsx"
wb.save(filepath)
print(f"Saved: {filepath}")
