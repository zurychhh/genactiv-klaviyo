#!/usr/bin/env python3
"""
Generuje raport spójności z planem napraw: 4 arkusze
1. Niespójności
2. Co naprawimy przez MCP
3. Single Source of Truth
4. Konwencja UTM
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
HEADER_FILL = PatternFill(start_color="1A3B5D", end_color="1A3B5D", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

RED_FONT = Font(name="Calibri", bold=True, size=10, color="9C0006")
YELLOW_FONT = Font(name="Calibri", bold=True, size=10, color="9C6500")
GREEN_FONT = Font(name="Calibri", bold=True, size=10, color="006100")
BLUE_FONT = Font(name="Calibri", bold=True, size=10, color="1A3B5D")

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1A3B5D")
SUBTITLE_FONT = Font(name="Calibri", bold=False, size=11, color="4472C4")
BOLD = Font(name="Calibri", bold=True, size=10)
NORMAL = Font(name="Calibri", size=10)
MONO = Font(name="Consolas", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_borders(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def auto_width(ws, max_col, min_width=10, max_width=45):
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    longest_line = max(len(line) for line in lines)
                    max_len = max(max_len, min(longest_line + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def write_data_row(ws, row, data, wrap=True):
    """Write a list of values to a row with standard formatting."""
    for j, val in enumerate(data):
        cell = ws.cell(row=row, column=j + 1, value=val)
        cell.font = NORMAL
        cell.alignment = Alignment(vertical="center", wrap_text=wrap)


# ============================================================
# SHEET 0: JAK CZYTAC RAPORT
# ============================================================
ws0 = wb.active
ws0.title = "Jak czytac raport"
ws0.sheet_properties.tabColor = "1A3B5D"

# Title
ws0.merge_cells("A1:H1")
ws0.cell(row=1, column=1, value="AUDYT DANYCH MARKETINGOWYCH — WYNIKI I PROPOZYCJE USPRAWNIEN").font = Font(name="Calibri", bold=True, size=20, color="1A3B5D")
ws0.row_dimensions[1].height = 36

ws0.merge_cells("A2:H2")
ws0.cell(row=2, column=1, value="GenActiv.pl — Klaviyo + Shopify + GA4 + Google Ads + Meta Ads").font = Font(name="Calibri", size=13, color="4472C4")

ws0.merge_cells("A3:H3")
ws0.cell(row=3, column=1, value="Wygenerowano: 02.03.2026 | Dane za okres: 20-26.02.2026 (weryfikacja) + biezacy stan konfiguracji").font = Font(name="Calibri", size=10, color="7F8C8D")

# What is this report
r = 5
ws0.merge_cells(f"A{r}:H{r}")
ws0.cell(row=r, column=1, value="CZYM JEST TEN RAPORT?").font = Font(name="Calibri", bold=True, size=13, color="1A3B5D")
ws0.cell(row=r, column=1).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

intro_texts = [
    "Ten raport prezentuje wyniki audytu spojnosci danych miedzy narzedziami marketingowymi GenActiv.pl.",
    "Porownalismy dane z 5 zrodel: Shopify (zamowienia), GA4 (sesje/ruch), Klaviyo (email), Google Ads i Meta Ads.",
    "Zidentyfikowalismy 13 obszarow, w ktorych dane z roznych narzedzi sie rozjezdzaja — co utrudnia trafne raportowanie.",
    "Dla kazdego obszaru przygotowalismy propozycje usprawnienia — do wspolnej oceny i decyzji, co wdrozyc.",
]
for i, txt in enumerate(intro_texts):
    rr = r + 1 + i
    ws0.merge_cells(f"A{rr}:H{rr}")
    ws0.cell(row=rr, column=1, value=txt).font = NORMAL
    ws0.row_dimensions[rr].height = 20

# Guide to sheets
r = r + len(intro_texts) + 2
ws0.merge_cells(f"A{r}:H{r}")
ws0.cell(row=r, column=1, value="PRZEWODNIK PO ARKUSZACH").font = Font(name="Calibri", bold=True, size=13, color="1A3B5D")
ws0.cell(row=r, column=1).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

guide_headers = ["Arkusz", "Co zawiera", "Dla kogo", "Jak czytac"]
r += 1
for i, h in enumerate(guide_headers, 1):
    ws0.cell(row=r, column=i, value=h)
style_header_row(ws0, r, len(guide_headers))
ws0.row_dimensions[r].height = 28

guide_data = [
    [
        "1. Niespojnosci",
        "Lista 13 zidentyfikowanych obszarow, w ktorych dane z roznych narzedzi sie rozjezdzaja — wraz z opisem wplywu na raportowanie.",
        "Wszyscy interesariusze",
        "Warto zaczac tutaj. Kolumna 'Priorytet' sugeruje pilnosc:\n"
        "KRYTYCZNY (czerwony) = dane sa trudne do uzycia\n"
        "WYSOKI (zolty) = dane sa fragmentaryczne\n"
        "SREDNI (niebieski) = dane sa niepelne, ale uzywalne\n"
        "NISKI (zielony) = drobne niedogodnosci",
    ],
    [
        "2. Propozycje napraw",
        "Dla kazdego obszaru — proponowane narzedzie, metoda i zakres dzialania. Wskazujemy tez, co mozna zautomatyzowac, a co wymagaloby recznej pracy.",
        "Zespol techniczny / marketing ops",
        "Kolumna 'Automatyzowalne?' wskazuje:\n"
        "TAK (zielony) = mozliwe do wdrozenia przez API/MCP\n"
        "CZESCIOWO (zolty) = czesc auto, czesc recznie\n"
        "NIE (czerwony) = wymagaloby recznej pracy w panelu narzedzia",
    ],
    [
        "3. Single Source of Truth",
        "Propozycja mapowania: dla kazdej metryki biznesowej (sesje, przychod, email, reklamy) — ktore narzedzie mogloby pelnic role zrodla prawdy i dlaczego.",
        "Osoby raportujace / analityk",
        "Moze sluzyc jako referencja przy tworzeniu raportow — np. przychod najlepiej weryfikowac w Shopify, ruch na stronie w GA4.\n"
        "Na dole znajduje sie diagram proponowanej architektury danych.",
    ],
    [
        "4. Propozycja konwencji UTM",
        "Proponowany standard tagowania linkow UTM dla kazdej platformy: Klaviyo, Meta Ads, Instagram organic, Google Ads, Bing, TikTok, YouTube, QR kody.",
        "Kazdy kto tworzy kampanie lub posty z linkami",
        "Proponujemy przyjac te zasady jako wspolny standard. Jesli zdecydujecie sie go wdrozyc, kazdzy link w kampanii / poscie / bio powinien miec UTM wedlug tej tabeli.\n"
        "Na dole 9 proponowanych zasad (np. zawsze lowercase, nigdy spacji).",
    ],
]

for i, d in enumerate(guide_data):
    rr = r + 1 + i
    for j, val in enumerate(d):
        cell = ws0.cell(row=rr, column=j + 1, value=val)
        cell.font = NORMAL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws0.row_dimensions[rr].height = 100
    # Color sheet name
    ws0.cell(row=rr, column=1).font = Font(name="Calibri", bold=True, size=10, color="1A3B5D")

apply_borders(ws0, r, r + len(guide_data), len(guide_headers))

# Legend of colors
r_legend = r + len(guide_data) + 2
ws0.merge_cells(f"A{r_legend}:H{r_legend}")
ws0.cell(row=r_legend, column=1, value="LEGENDA KOLOROW").font = Font(name="Calibri", bold=True, size=13, color="1A3B5D")
ws0.cell(row=r_legend, column=1).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

legend_items = [
    ("KRYTYCZNY / GAP / NIE", RED_FILL, RED_FONT, "Obszar o duzym wplywie na dane. Proponujemy priorytetowe zajecie sie tym."),
    ("WYSOKI / WARN / CZESCIOWO", YELLOW_FILL, YELLOW_FONT, "Powoduje fragmentacje danych. Warto rozwazyc naprawe w ciagu 1-2 tygodni."),
    ("SREDNI / INFO", BLUE_FILL, BLUE_FONT, "Utrudnia analize, ale dane sa uzywalne. Mozna zaplanowac na pozniej."),
    ("NISKI / OK / TAK", GREEN_FILL, GREEN_FONT, "Drobna kwestia lub potwierdzenie poprawnosci. Niska pilnosc."),
]

for i, (label, fill, font, desc) in enumerate(legend_items):
    rr = r_legend + 1 + i
    cell_label = ws0.cell(row=rr, column=1, value=label)
    cell_label.fill = fill
    cell_label.font = font
    cell_label.alignment = Alignment(horizontal="center", vertical="center")
    ws0.merge_cells(f"B{rr}:D{rr}")
    ws0.cell(row=rr, column=2, value=desc).font = NORMAL
    ws0.cell(row=rr, column=2).alignment = Alignment(vertical="center", wrap_text=True)

apply_borders(ws0, r_legend, r_legend + len(legend_items), 4)

# Key numbers summary
r_key = r_legend + len(legend_items) + 2
ws0.merge_cells(f"A{r_key}:H{r_key}")
ws0.cell(row=r_key, column=1, value="KLUCZOWE LICZBY Z AUDYTU").font = Font(name="Calibri", bold=True, size=13, color="1A3B5D")
ws0.cell(row=r_key, column=1).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

key_numbers = [
    ("13", "zidentyfikowanych niespojnosci w danych"),
    ("2", "problemy KRYTYCZNE (niewidoczne kampanie email, rozbity utm_source)"),
    ("5", "problemow WYSOKICH (case mismatch, brak UTM, Przelewy24)"),
    ("6/13", "problemow mozna naprawic automatycznie przez MCP (API)"),
    ("5/13", "problemow wymaga recznej interwencji w panelach narzedzi"),
    ("2/13", "problemy nie do naprawienia (akceptujemy ograniczenia)"),
    ("~38%", "roznica w przychodach miedzy Shopify (77.6k) a GA4 (48.4k) — z powodu niespojnosci"),
    ("79%", "zamowien w Shopify bez kompletnych parametrow UTM w last visit"),
]

for i, (num, desc) in enumerate(key_numbers):
    rr = r_key + 1 + i
    cell_num = ws0.cell(row=rr, column=1, value=num)
    cell_num.font = Font(name="Calibri", bold=True, size=14, color="EF3340")
    cell_num.alignment = Alignment(horizontal="center", vertical="center")
    ws0.merge_cells(f"B{rr}:D{rr}")
    ws0.cell(row=rr, column=2, value=desc).font = NORMAL
    ws0.cell(row=rr, column=2).alignment = Alignment(vertical="center")

apply_borders(ws0, r_key, r_key + len(key_numbers), 4)

# Column widths
ws0.column_dimensions["A"].width = 22
ws0.column_dimensions["B"].width = 45
ws0.column_dimensions["C"].width = 30
ws0.column_dimensions["D"].width = 45

ws0.sheet_view.showGridLines = False


# ============================================================
# SHEET 1: NIESPOOJNOSCI
# ============================================================
ws1 = wb.create_sheet("1. Niespojnosci")
ws1.sheet_properties.tabColor = "EF3340"

ws1.merge_cells("A1:F1")
ws1.cell(row=1, column=1, value="LISTA NIESPO\u0301JNOS\u0301CI W RAPORTOWANIU \u2014 GenActiv.pl").font = TITLE_FONT
ws1.merge_cells("A2:F2")
ws1.cell(row=2, column=1, value="Stan na: 02.03.2026").font = SUBTITLE_FONT

headers1 = ["#", "Kategoria", "Problem", "Wp\u0142yw na raportowanie", "Dotknie\u0327te narze\u0327dzia", "Priorytet"]
row = 4
for i, h in enumerate(headers1, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, len(headers1))
ws1.row_dimensions[row].height = 30

data1 = [
    [1, "UTM source email",
     "utm_source='newsletter' w starych kampaniach vs 'Klaviyo' w nowych",
     "Email traffic rozbity na 2 \u017ar\u00f3d\u0142a w GA4 i Shopify. Nie da si\u0119 raportowa\u0107 \u0142\u0105cznie email performance.",
     "GA4, Shopify, Klaviyo", "KRYTYCZNY"],
    [2, "UTM campaign case mismatch",
     "JESIENGENACTIV vs JesienGenactiv \u2014 ta sama kampania, r\u00f3\u017cny case",
     "GA4 traktuje jako 2 osobne kampanie. 66 vs 22 sesje zamiast \u0142\u0105cznych 88.",
     "GA4, Shopify", "WYSOKI"],
    [3, "Meta Ads source case",
     "instagram vs Instagram, facebook vs Facebook \u2014 r\u00f3\u017cny case w UTMach",
     "Fragmentacja: 3+ wierszy w GA4 zamiast jednego. Raporty social media niepe\u0142ne.",
     "GA4, Meta Ads Manager", "WYSOKI"],
    [4, "Meta Ads medium niesp\u00f3jny",
     "paidsocial vs cpc vs social vs Bio \u2014 4 r\u00f3\u017cne medium dla podobnych \u017ar\u00f3de\u0142",
     "Nie da si\u0119 filtrowa\u0107 paid vs organic social. Default Channel Group w GA4 \u017ale kategoryzuje.",
     "GA4, Meta Ads", "WYSOKI"],
    [5, "Brak UTM w linkach social organic",
     "l.instagram.com/referral (309 sesji), m.facebook.com/referral (53 sesji) \u2014 zero UTM\u00f3w",
     "362 sesje social wpadaj\u0105 jako 'referral' zamiast 'social'. 17 zam\u00f3wie\u0144 bez atrybucji.",
     "GA4, Shopify, Instagram, Facebook", "WYSOKI"],
    [6, "Lutowe kampanie Klaviyo niewidoczne w GA4",
     "Kampanie A2, Junior, Follow-up z lutego 2026 \u2014 0 sesji w GA4",
     "Kompletna utrata danych o efektywno\u015bci najnowszych kampanii email.",
     "Klaviyo, GA4", "KRYTYCZNY"],
    [7, "Przelewy24 przerywa sesj\u0119 GA4",
     "go.przelewy24.pl nie jest na li\u015bcie Unwanted Referrals w GA4",
     "Po powrocie z bramki p\u0142atno\u015bci GA4 tworzy now\u0105 sesj\u0119 'referral' zamiast kontynuowa\u0107 oryginaln\u0105.",
     "GA4", "WYSOKI"],
    [8, "Brak utm_content w Klaviyo",
     "17/18 zam\u00f3wie\u0144 email ma utm_content=(empty)",
     "Nie da si\u0119 A/B testowa\u0107 szablon\u00f3w email ani analizowa\u0107 kt\u00f3re CTA konwertuj\u0105.",
     "Klaviyo, GA4", "\u015aREDNI"],
    [9, "Gmail App stripuje UTMy",
     "android-app://com.google.android.gm/ \u2014 4 zam\u00f3wienia bez UTM",
     "Email na Androidzie traci atrybucj\u0119. Nie do ko\u0144ca naprawialne.",
     "Email clients", "NISKI"],
    [10, "Kampania '3528' niezidentyfikowana",
     "301 sesji, 11 konwersji, 3021 PLN z nieznanej kampanii '3528'",
     "3k PLN przychodu bez atrybucji do \u017ar\u00f3d\u0142a. Prawdopodobnie Meta Ads lub inny pixel.",
     "GA4, ?", "\u015aREDNI"],
    [11, "Consent Mode \u2014 15% Unassigned",
     "1491 sesji (15%) w GA4 jako Unassigned z powodu braku zgody analytics_storage",
     "GA4 nie widzi 15% ruchu. Por\u00f3wnanie z Shopify zawsze b\u0119dzie rozbie\u017cne.",
     "GA4, Pandectes, GTM", "\u015aREDNI"],
    [12, "Shopify Analytics API niedost\u0119pne",
     "ShopifyQL nie dzia\u0142a przez custom app token. Brak danych o sesjach z Shopify.",
     "Nie mamy session-level data z Shopify do cross-referencji. Tylko order-level journey.",
     "Shopify", "NISKI"],
    [13, "admin.shopify.com jako referral",
     "39 sesji, 1 konwersja z admin.shopify.com/referral",
     "Wewn\u0119trzny ruch Shopify admin mieszany z customer traffic.",
     "GA4", "NISKI"],
]

PRIORITY_STYLES = {
    "KRYTYCZNY": (RED_FILL, RED_FONT),
    "WYSOKI": (YELLOW_FILL, YELLOW_FONT),
    "\u015aREDNI": (BLUE_FILL, BLUE_FONT),
    "NISKI": (GREEN_FILL, GREEN_FONT),
}

for i, d in enumerate(data1):
    r = row + 1 + i
    write_data_row(ws1, r, d)
    # Center the # column
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    # Color priority
    prio_cell = ws1.cell(row=r, column=6)
    prio = str(prio_cell.value)
    if prio in PRIORITY_STYLES:
        fill, font = PRIORITY_STYLES[prio]
        prio_cell.fill = fill
        prio_cell.font = font
        prio_cell.alignment = Alignment(horizontal="center", vertical="center")

apply_borders(ws1, row, row + len(data1), len(headers1))
auto_width(ws1, len(headers1))
# Override specific column widths for readability
ws1.column_dimensions["A"].width = 5
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 45
ws1.column_dimensions["D"].width = 45
ws1.column_dimensions["E"].width = 30
ws1.column_dimensions["F"].width = 14
ws1.freeze_panes = "B5"


# ============================================================
# SHEET 2: CO NAPRAWIMY PRZEZ MCP
# ============================================================
ws2 = wb.create_sheet("2. Propozycje napraw")
ws2.sheet_properties.tabColor = "27AE60"

ws2.merge_cells("A1:G1")
ws2.cell(row=1, column=1, value="PROPOZYCJE NAPRAW \u2014 Co mo\u017cna zautomatyzowa\u0107 vs r\u0119czne").font = TITLE_FONT

headers2 = ["#", "Niespo\u0301jnos\u0301c\u0301 (ref)", "Narze\u0327dzie MCP", "Metoda/Tool",
            "Co dok\u0142adnie zrobimy", "Automatyzowalne?", "Status"]
r2 = 3
for i, h in enumerate(headers2, 1):
    ws2.cell(row=r2, column=i, value=h)
style_header_row(ws2, r2, len(headers2))
ws2.row_dimensions[r2].height = 35

data2 = [
    [1, "#1 UTM source email", "Klaviyo MCP",
     "klaviyo_create_email_template + r\u0119czna aktualizacja kampanii",
     "Zaktualizowa\u0107 szablony HTML: zamieni\u0107 utm_source=newsletter na utm_source=klaviyo we wszystkich linkach. Nowe kampanie ustawia\u0107 z custom_tracking_params.",
     "CZ\u0118\u015aCIOWO \u2014 szablony tak, stare kampanie nie (already sent)", "DO ZROBIENIA"],
    [2, "#2 Campaign case", "Klaviyo MCP",
     "Nowe kampanie z lowercase naming convention",
     "Ustali\u0107 konwencj\u0119: zawsze lowercase campaign name. Wdro\u017cy\u0107 w przysz\u0142ych kampaniach.",
     "TAK \u2014 nowe kampanie", "DO ZROBIENIA"],
    [3, "#3 Meta source case", "Meta Ads MCP",
     "update_campaign / update_adset \u2014 URL parameters",
     "Sprawdzi\u0107 i zaktualizowa\u0107 URL Parameters w aktywnych kampaniach: wymusi\u0107 lowercase utm_source=instagram lub facebook.",
     "TAK", "DO ZROBIENIA"],
    [4, "#4 Meta medium", "Meta Ads MCP",
     "update_campaign \u2014 URL parameters",
     "Ustandaryzowa\u0107: paid=paidsocial, organic=social. Zaktualizowa\u0107 w aktywnych kampaniach.",
     "TAK", "DO ZROBIENIA"],
    [5, "#5 Brak UTM social organic", "\u2014",
     "R\u0118CZNIE",
     "Doda\u0107 UTMy do: (a) link w bio Instagram, (b) link w bio Facebook, (c) linki w postach organicznych. Wz\u00f3r: utm_source=instagram&utm_medium=social&utm_campaign=bio_link",
     "NIE \u2014 wymaga r\u0119cznej zmiany w Instagram/Facebook", "DO ZROBIENIA"],
    [6, "#6 Lutowe kampanie", "Klaviyo MCP",
     "klaviyo_get_email_template + analiza HTML",
     "Pobra\u0107 szablony lutowych kampanii, przeanalizowa\u0107 HTML pod k\u0105tem link\u00f3w bez UTM. Naprawi\u0107 i re-u\u017cy\u0107 w follow-upach.",
     "TAK \u2014 analiza i fix szablon\u00f3w", "DO ZROBIENIA"],
    [7, "#7 Przelewy24", "\u2014",
     "R\u0118CZNIE w GA4 Admin",
     "Wej\u015b\u0107 w GA4 Admin \u2192 Data Streams \u2192 Configure tag settings \u2192 List unwanted referrals \u2192 doda\u0107 'go.przelewy24.pl'",
     "NIE \u2014 GA4 Admin nie ma API do tego", "DO ZROBIENIA"],
    [8, "#8 Brak utm_content", "Klaviyo MCP",
     "custom_tracking_params w nowych kampaniach",
     "Dodawa\u0107 utm_content w szablonach (np. variant_a, header_cta, product_card). Wdro\u017cy\u0107 jako standard.",
     "TAK \u2014 nowe kampanie", "DO ZROBIENIA"],
    [9, "#9 Gmail App", "\u2014",
     "NIE DA SI\u0118 NAPRAWI\u0106",
     "Gmail App na Androidzie stripuje UTMy z link\u00f3w email. Workaround: u\u017cywa\u0107 redirect URLs (np. przez Klaviyo tracking) kt\u00f3re koduj\u0105 UTM w path.",
     "NIE", "AKCEPTUJEMY"],
    [10, "#10 Kampania 3528", "GA4 MCP + Meta Ads MCP",
     "run_report + get_campaigns",
     "Przeszuka\u0107 Meta Ads i Google Ads po ID '3528'. Sprawdzi\u0107 czy to campaign ID z pixela Meta.",
     "TAK \u2014 research", "DO ZROBIENIA"],
    [11, "#11 Consent Unassigned", "\u2014",
     "R\u0118CZNIE w Pandectes",
     "Zweryfikowa\u0107 konfiguracj\u0119 bannera: czy analytics_storage jest mapowane prawid\u0142owo. Rozwa\u017cy\u0107 Google Consent Mode modeling.",
     "NIE \u2014 Pandectes dashboard", "DO ZROBIENIA"],
    [12, "#12 Shopify Analytics", "\u2014",
     "OGRANICZONE",
     "ShopifyQL wymaga Shopify Plus lub embedded app. Workaround: u\u017cywamy GA4 jako source of truth dla sesji.",
     "NIE", "AKCEPTUJEMY"],
    [13, "#13 admin.shopify.com", "\u2014",
     "R\u0118CZNIE w GA4 Admin",
     "Doda\u0107 admin.shopify.com do Unwanted Referrals w GA4 (razem z Przelewy24).",
     "NIE \u2014 GA4 Admin", "DO ZROBIENIA"],
]

AUTO_STYLES = {
    "TAK": (GREEN_FILL, GREEN_FONT),
    "CZ\u0118\u015aCIOWO": (YELLOW_FILL, YELLOW_FONT),
    "NIE": (RED_FILL, RED_FONT),
}

STATUS_STYLES = {
    "DO ZROBIENIA": (YELLOW_FILL, YELLOW_FONT),
    "AKCEPTUJEMY": (BLUE_FILL, BLUE_FONT),
}

for i, d in enumerate(data2):
    r = r2 + 1 + i
    write_data_row(ws2, r, d)
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Color "Automatyzowalne?" column (6)
    auto_cell = ws2.cell(row=r, column=6)
    auto_val = str(auto_cell.value)
    for key, (fill, font) in AUTO_STYLES.items():
        if auto_val.startswith(key):
            auto_cell.fill = fill
            auto_cell.font = font
            break

    # Color "Status" column (7)
    status_cell = ws2.cell(row=r, column=7)
    status_val = str(status_cell.value)
    if status_val in STATUS_STYLES:
        fill, font = STATUS_STYLES[status_val]
        status_cell.fill = fill
        status_cell.font = font
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

apply_borders(ws2, r2, r2 + len(data2), len(headers2))
auto_width(ws2, len(headers2))
ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 25
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 35
ws2.column_dimensions["E"].width = 45
ws2.column_dimensions["F"].width = 30
ws2.column_dimensions["G"].width = 16
ws2.freeze_panes = "B4"


# ============================================================
# SHEET 3: SINGLE SOURCE OF TRUTH
# ============================================================
ws3 = wb.create_sheet("3. Single Source of Truth")
ws3.sheet_properties.tabColor = "4472C4"

ws3.merge_cells("A1:E1")
ws3.cell(row=1, column=1, value="ARCHITEKTURA DANYCH \u2014 Single Source of Truth").font = TITLE_FONT

headers3 = ["Metryka", "Source of Truth", "Dlaczego", "Alternatywne \u017ar\u00f3d\u0142o", "Uwagi"]
r3 = 3
for i, h in enumerate(headers3, 1):
    ws3.cell(row=r3, column=i, value=h)
style_header_row(ws3, r3, len(headers3))
ws3.row_dimensions[r3].height = 30

data3 = [
    ["Sesje / ruch na stronie", "GA4",
     "Jedyne narz\u0119dzie z session-level tracking. Shopify API nie daje sesji.",
     "Shopify Admin (manual)",
     "Pami\u0119taj: GA4 nie widzi ~15% ruchu (consent)"],
    ["\u0179r\u00f3d\u0142a ruchu (acquisition)", "GA4",
     "Najdok\u0142adniejsze UTM + auto-tagging (gclid). Shopify daje tylko order-level.",
     "Shopify customer journey",
     "Cross-reference z Shopify journey dla email/social"],
    ["Zam\u00f3wienia i przych\u00f3d", "SHOPIFY",
     "Source of truth dla transakcji. GA4 traci ~34% konwersji.",
     "GA4 e-commerce",
     "Shopify = 72k PLN vs GA4 = 48k PLN za ten sam okres"],
    ["Email performance (opens, clicks)", "KLAVIYO",
     "Jedyne narz\u0119dzie z email engagement metrics.",
     "GA4 (po klikni\u0119ciu)",
     "GA4 widzi tylko sesje po klikni\u0119ciu, nie opens"],
    ["Email attribution (sprzeda\u017c)", "SHOPIFY + KLAVIYO",
     "Shopify journey daje last-click, Klaviyo daje attributed revenue.",
     "GA4",
     "Por\u00f3wnuj oba \u2014 Klaviyo zawy\u017ca, GA4 zani\u017ca"],
    ["Google Ads performance", "GOOGLE ADS",
     "Source of truth: impressions, clicks, cost, CPC. GA4 dla on-site behavior.",
     "GA4",
     "Google Ads conversions mog\u0105 si\u0119 r\u00f3\u017cni\u0107 od GA4 (conversion window)"],
    ["Meta Ads performance", "META ADS",
     "Source of truth: reach, impressions, clicks, cost.",
     "GA4",
     "Meta attributed conversions vs GA4 last-click \u2014 zawsze rozbie\u017cne"],
    ["SEO traffic", "GA4",
     "google/organic sessions, landing pages, behavior.",
     "Google Search Console",
     "GSC dla impressions/CTR, GA4 dla on-site"],
    ["Customer data / profiles", "SHOPIFY",
     "Master customer database. Klaviyo syncs from Shopify.",
     "Klaviyo profiles",
     "Shopify = transactional truth, Klaviyo = marketing engagement"],
    ["Product catalog", "SHOPIFY",
     "Source of truth dla produkt\u00f3w, cen, stan\u00f3w.",
     "Klaviyo catalog",
     "Klaviyo catalog syncs from Shopify"],
]

# Color the Source of Truth column with tool-specific colors
SOT_COLORS = {
    "GA4": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "SHOPIFY": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "KLAVIYO": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "GOOGLE ADS": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "META ADS": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "SHOPIFY + KLAVIYO": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
}

for i, d in enumerate(data3):
    r = r3 + 1 + i
    write_data_row(ws3, r, d)
    sot_cell = ws3.cell(row=r, column=2)
    sot_val = str(sot_cell.value)
    if sot_val in SOT_COLORS:
        sot_cell.fill = SOT_COLORS[sot_val]
        sot_cell.font = Font(name="Calibri", bold=True, size=10)
        sot_cell.alignment = Alignment(horizontal="center", vertical="center")

apply_borders(ws3, r3, r3 + len(data3), len(headers3))

# Dashboard Architecture diagram
diagram_start = r3 + len(data3) + 3
ws3.merge_cells(f"A{diagram_start}:E{diagram_start}")
ws3.cell(row=diagram_start, column=1, value="Dashboard Architecture").font = Font(
    name="Calibri", bold=True, size=12, color="1A3B5D")

diagram_text = (
    "\u250c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2510\n"
    "\u2502              SINGLE SOURCE OF TRUTH                  \u2502\n"
    "\u251c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u252c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u252c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u252c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u252c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2524\n"
    "\u2502  TRAFFIC \u2502  SALES   \u2502  EMAIL   \u2502 PAID ADS \u2502   SEO   \u2502\n"
    "\u2502   GA4    \u2502 SHOPIFY  \u2502 KLAVIYO  \u2502 GA/META  \u2502GA4+GSC  \u2502\n"
    "\u251c\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2534\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2534\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2534\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2534\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2524\n"
    "\u2502         DASHBOARD OPERACYJNY (Excel/Script)          \u2502\n"
    "\u2502    Pobiera dane z ka\u017cdego narz\u0119dzia przez MCP        \u2502\n"
    "\u2514\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2518"
)

diagram_row = diagram_start + 1
ws3.merge_cells(f"A{diagram_row}:E{diagram_row + 8}")
cell = ws3.cell(row=diagram_row, column=1, value=diagram_text)
cell.font = MONO
cell.alignment = Alignment(vertical="top", wrap_text=True)
# Set row height to fit the diagram
for rr in range(diagram_row, diagram_row + 9):
    ws3.row_dimensions[rr].height = 16

auto_width(ws3, len(headers3))
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 45
ws3.column_dimensions["D"].width = 25
ws3.column_dimensions["E"].width = 45
ws3.freeze_panes = "B4"


# ============================================================
# SHEET 4: KONWENCJA UTM
# ============================================================
ws4 = wb.create_sheet("4. Propozycja konwencji UTM")
ws4.sheet_properties.tabColor = "0066CC"

ws4.merge_cells("A1:G1")
ws4.cell(row=1, column=1, value="STANDARD UTM \u2014 Proponowana konwencja").font = TITLE_FONT

headers4 = ["Platforma", "utm_source", "utm_medium", "utm_campaign",
            "utm_content", "utm_term", "Przyk\u0142ad pe\u0142nego URL"]
r4 = 3
for i, h in enumerate(headers4, 1):
    ws4.cell(row=r4, column=i, value=h)
style_header_row(ws4, r4, len(headers4))
ws4.row_dimensions[r4].height = 30

data4 = [
    ["Klaviyo (email kampanie)", "klaviyo", "email", "{campaign_name} lowercase",
     "{template_variant}", "{product_keywords}",
     "?utm_source=klaviyo&utm_medium=email&utm_campaign=colostrum_a2&utm_content=variant_a&utm_term=colostrum"],
    ["Klaviyo (flows)", "klaviyo", "email", "flow_{flow_name}",
     "{step_name}", "{product}",
     "?utm_source=klaviyo&utm_medium=email&utm_campaign=flow_abandoned_cart&utm_content=reminder_1"],
    ["Meta Ads (Instagram)", "instagram", "paidsocial", "{campaign_name}",
     "{ad_name}", "{targeting}",
     "?utm_source=instagram&utm_medium=paidsocial&utm_campaign=junior_winter_2026&utm_content=animated_ad"],
    ["Meta Ads (Facebook)", "facebook", "paidsocial", "{campaign_name}",
     "{ad_name}", "{targeting}",
     "?utm_source=facebook&utm_medium=paidsocial&utm_campaign=junior_winter_2026&utm_content=carousel_1"],
    ["Instagram organic (bio)", "instagram", "social", "bio_link",
     "(opcjonalnie)", "(opcjonalnie)",
     "?utm_source=instagram&utm_medium=social&utm_campaign=bio_link"],
    ["Instagram organic (stories)", "instagram", "social", "{post_topic}",
     "stories", "(opcjonalnie)",
     "?utm_source=instagram&utm_medium=social&utm_campaign=colostrum_a2&utm_content=stories"],
    ["Facebook organic", "facebook", "social", "{post_topic}",
     "{post_type}", "(opcjonalnie)",
     "?utm_source=facebook&utm_medium=social&utm_campaign=colostrum_a2&utm_content=post"],
    ["Google Ads", "(auto \u2014 gclid)", "(auto)", "(auto)",
     "(auto)", "(auto)",
     "Nie ustawia\u0107 UTM! Auto-tagging przez gclid."],
    ["Bing Ads (Sembot)", "bing", "cpc", "{campaign_name}",
     "{ad_group}", "{keywords}",
     "Zarz\u0105dzane przez Sembot \u2014 zweryfikowa\u0107 sp\u00f3jno\u015b\u0107"],
    ["TikTok Ads", "tiktok", "paidsocial", "{campaign_name}",
     "{ad_name}", "(opcjonalnie)",
     "?utm_source=tiktok&utm_medium=paidsocial&utm_campaign=junior_winter_2026"],
    ["YouTube organic", "youtube", "social", "{video_topic}",
     "{video_type}", "(opcjonalnie)",
     "?utm_source=youtube&utm_medium=social&utm_campaign=colostrum_shorts"],
    ["QR kod / ulotka", "qrkod", "offline", "{material_name}",
     "{location}", "(opcjonalnie)",
     "?utm_source=qrkod&utm_medium=offline&utm_campaign=probki_apteka"],
]

# Platform-specific row coloring (subtle alternating)
PLATFORM_FILLS = {
    "Klaviyo": PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid"),
    "Meta": PatternFill(start_color="FFF7F0", end_color="FFF7F0", fill_type="solid"),
    "Instagram": PatternFill(start_color="FFF7F0", end_color="FFF7F0", fill_type="solid"),
    "Facebook": PatternFill(start_color="FFF7F0", end_color="FFF7F0", fill_type="solid"),
    "Google": PatternFill(start_color="F0F7F0", end_color="F0F7F0", fill_type="solid"),
    "Bing": PatternFill(start_color="F7F7F0", end_color="F7F7F0", fill_type="solid"),
    "TikTok": PatternFill(start_color="F0F0F7", end_color="F0F0F7", fill_type="solid"),
    "YouTube": PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid"),
    "QR": PatternFill(start_color="F7F0F7", end_color="F7F0F7", fill_type="solid"),
}

for i, d in enumerate(data4):
    r = r4 + 1 + i
    write_data_row(ws4, r, d)
    # Apply subtle platform fill
    platform = str(d[0])
    for key, fill in PLATFORM_FILLS.items():
        if key in platform:
            for c in range(1, len(headers4) + 1):
                ws4.cell(row=r, column=c).fill = fill
            break

apply_borders(ws4, r4, r4 + len(data4), len(headers4))

# ZASADY section
zasady_start = r4 + len(data4) + 2
ws4.merge_cells(f"A{zasady_start}:G{zasady_start}")
zasady_title_cell = ws4.cell(row=zasady_start, column=1, value="PROPONOWANE ZASADY UTM:")
zasady_title_cell.font = Font(name="Calibri", bold=True, size=12, color="1A3B5D")
zasady_title_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

zasady = [
    "1. Wszystko lowercase (nigdy Uppercase w source/medium/campaign)",
    "2. Spacje zamieniane na _ (underscore)",
    "3. utm_source = nazwa platformy (klaviyo, instagram, facebook, tiktok, bing, youtube, qrkod)",
    "4. utm_medium = typ ruchu (email, paidsocial, social, cpc, offline)",
    "5. utm_campaign = nazwa kampanii (sp\u00f3jna mi\u0119dzy narz\u0119dziami)",
    "6. utm_content = wariant reklamy/szablonu (do A/B test\u00f3w)",
    "7. utm_term = s\u0142owa kluczowe lub produkty",
    "8. Google Ads: NIE dodawa\u0107 UTM \u2014 u\u017cywa\u0107 auto-tagging (gclid)",
    "9. Zawsze testowa\u0107 linki przed wysy\u0142k\u0105 kampanii",
]

for idx, z in enumerate(zasady):
    rr = zasady_start + 1 + idx
    ws4.merge_cells(f"A{rr}:G{rr}")
    cell = ws4.cell(row=rr, column=1, value=z)
    cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(vertical="center", indent=1)
    # Alternate light grey background
    if idx % 2 == 0:
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Add thin borders to zasady
for rr in range(zasady_start, zasady_start + 1 + len(zasady)):
    for c in range(1, len(headers4) + 1):
        ws4.cell(row=rr, column=c).border = THIN_BORDER

auto_width(ws4, len(headers4))
ws4.column_dimensions["A"].width = 25
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 16
ws4.column_dimensions["D"].width = 25
ws4.column_dimensions["E"].width = 20
ws4.column_dimensions["F"].width = 18
ws4.column_dimensions["G"].width = 45
ws4.freeze_panes = "B4"


# ============================================================
# SAVE
# ============================================================
filepath = "/Users/user/projects/genactiv-klaviyo/reports/raport_spojnosci_plan_napraw.xlsx"
wb.save(filepath)
print(f"Saved: {filepath}")
print(f"Sheets: {wb.sheetnames}")
print("Done!")
